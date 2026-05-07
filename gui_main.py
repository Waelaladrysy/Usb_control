# -*- coding: utf-8 -*-
"""
USB Security System Pro - Auto-Block Edition + Actual Blocking
✅ التحديثات:
منع فعلي عبر Registry
مراقبة مستمرة في الخلفية
جدول auto_blocked منفصل
أزرار Refresh في كل القوائم
نظام مصادقة للعمليات الحساسة
نظام إشعارات باستخدام plyer
System Tray — الإخفاء بدلاً من الإغلاق عند ✕
Mutex — منع تشغيل أكثر من نسخة واحدة
"""
import threading
import time
import wmi
import tkinter as tk
import subprocess
from tkinter import ttk, messagebox, scrolledtext, simpledialog
import os
import sys
from datetime import datetime
import hashlib
import sqlite3
import winreg
import re
import ctypes

# ✅ استيراد plyer بشكل آمن
try:
    from plyer import notification
    PLYER_AVAILABLE = True
except ImportError:
    PLYER_AVAILABLE = False
    notification = None

# ✅ استيراد pystray + PIL للـ System Tray
try:
    import pystray
    from PIL import Image, ImageDraw
    TRAY_AVAILABLE = True
except ImportError:
    print("⚠️ pystray/Pillow not found — attempting auto-install...")
    try:
        import subprocess as _sp
        _sp.check_call(
            [sys.executable, "-m", "pip", "install", "pystray", "Pillow", "--quiet"],
            stdout=_sp.DEVNULL, stderr=_sp.DEVNULL
        )
        import pystray
        from PIL import Image, ImageDraw
        TRAY_AVAILABLE = True
        print("✅ pystray/Pillow installed successfully")
    except Exception as _e:
        print(f"❌ Could not install pystray/Pillow: {_e}")
        TRAY_AVAILABLE = False
        pystray = None

import pythoncom  # ✅ مهم جداً لـ WMI في الخلفية

# ==================== Mutex: منع تشغيل نسختين ====================
_MUTEX_NAME = "USBShieldProSingleInstance"
_mutex_handle = None

def _acquire_mutex() -> bool:
    """
    يُنشئ Mutex بالويندوز. يُرجع True إذا كانت هذه هي النسخة الأولى،
    False إذا كانت نسخة ثانية (يجب الخروج فوراً).
    """
    global _mutex_handle
    try:
        _mutex_handle = ctypes.windll.kernel32.CreateMutexW(None, True, _MUTEX_NAME)
        err = ctypes.windll.kernel32.GetLastError()
        return err != 183  # 183 = ERROR_ALREADY_EXISTS
    except Exception:
        return True  # في حالة الخطأ نسمح بالتشغيل

# ==================== إعدادات المسارات ====================
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(current_dir)
security_logic_path = os.path.join(project_root, "security_logic")
sys.path.insert(0, security_logic_path)

# ==================== بدء تلقائي مع الويندوز ====================
_APP_NAME    = "USBShieldPro"
_STARTUP_KEY = r"Software\Microsoft\Windows\CurrentVersion\Run"

def _get_exe_path() -> str:
    """
    يُرجع المسار الصحيح للبرنامج سواء كان .exe أو .py.
    عند التشغيل كـ PyInstaller exe: sys.executable هو UsbShield.exe
    عند التشغيل كـ Python script: نستخدم python + مسار السكريبت
    """
    if getattr(sys, 'frozen', False):
        # PyInstaller exe
        return os.path.abspath(sys.executable)
    else:
        # Python script — نستخدم pythonw.exe لتجنب نافذة CMD
        python_exe = sys.executable.replace('python.exe', 'pythonw.exe')
        script_path = os.path.abspath(__file__)
        return f'"{python_exe}" "{script_path}"'

def enable_startup():
    """تسجيل البرنامج ليعمل عند بدء تشغيل الويندوز — في الخلفية بدون نافذة"""
    try:
        exe = _get_exe_path()
        # إذا كان exe (مجمّع) نلفه بـ quotes، وإلا المسار مع pythonw مُعالَج سلفاً
        if getattr(sys, 'frozen', False):
            cmd = f'"{exe}" --startup'
        else:
            cmd = f'{exe} --startup'
        # ── كتابة في HKEY_CURRENT_USER (لا يحتاج Admin) ──────────────────
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _STARTUP_KEY,
                            0, winreg.KEY_SET_VALUE) as k:
            winreg.SetValueEx(k, _APP_NAME, 0, winreg.REG_SZ, cmd)
        print(f"✅ Startup registered: {cmd}")
        # ── التحقق الفوري من نجاح الكتابة ───────────────────────────────
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _STARTUP_KEY,
                            0, winreg.KEY_READ) as k:
            val, _ = winreg.QueryValueEx(k, _APP_NAME)
            if "--startup" not in val:
                raise ValueError(f"Startup value written without --startup flag: {val}")
        return True
    except Exception as e:
        print(f"⚠️ Could not enable startup: {e}")
        return False

def disable_startup():
    """إلغاء تسجيل البرنامج من بدء التشغيل"""
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _STARTUP_KEY,
                            0, winreg.KEY_SET_VALUE) as k:
            winreg.DeleteValue(k, _APP_NAME)
        return True
    except Exception:
        return False

def is_startup_enabled() -> bool:
    """هل البرنامج مسجّل في بدء التشغيل؟"""
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, _STARTUP_KEY,
                            0, winreg.KEY_READ) as k:
            val, _ = winreg.QueryValueEx(k, _APP_NAME)
            return bool(val)
    except Exception:
        return False

# ==================== أيقونة Tray (تُولَّد برمجياً) ====================
def _make_tray_icon(size: int = 64) -> "Image.Image":
    """
    يُولِّد أيقونة الـ tray بالكود — بدون ملف خارجي.
    درع أزرق داكن على خلفية سوداء.
    لا يعتمد على أي خط خارجي لضمان العمل في كل البيئات.
    """
    img  = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)

    # خلفية دائرية داكنة
    draw.ellipse([0, 0, size - 1, size - 1], fill=(11, 14, 20, 255))

    # شكل الدرع (6 نقاط)
    cx, cy = size / 2, size / 2
    s = size * 0.40
    shield = [
        (cx,          cy - s),
        (cx + s,      cy - s * 0.45),
        (cx + s,      cy + s * 0.25),
        (cx,          cy + s),
        (cx - s,      cy + s * 0.25),
        (cx - s,      cy - s * 0.45),
    ]
    draw.polygon(shield, fill=(0, 209, 255, 255))   # #00d1ff

    # مستطيل أبيض صغير في وسط الدرع (بديل عن حرف S)
    bar_w = size * 0.08
    bar_h = size * 0.35
    bx = cx - bar_w / 2
    by = cy - bar_h / 2
    draw.rectangle([bx, by, bx + bar_w, by + bar_h],
                   fill=(11, 14, 20, 255))

    return img

try:
    from database_manager import get_connection, log_event
    from auth_manager import login, change_password, change_username, verify_password
    from whitelist_manager import (
        add_to_whitelist, remove_from_whitelist,
        get_all_whitelist_devices, is_device_whitelisted,
        get_device_by_fingerprint as get_whitelist_device
    )
    from blacklist_manager import (
        add_to_blacklist, remove_from_blacklist,
        get_all_blacklist_devices, is_device_blacklisted,
        device_exists_in_blacklist, get_device_by_fingerprint as get_blacklist_device
    )
    from usb_monitor import USBMonitor
    from usb_blocker import USBBlocker
except ImportError as e:
    messagebox.showerror("Error", f"Failed to import modules:\n{e}")
    sys.exit(1)

# ==================== لوحة الألوان الاحترافية ====================
class Colors:
    BG_DARK = '#0b0e14'
    BG_CARD = '#151921'
    SIDEBAR = '#10141d'
    PRIMARY = '#00d1ff'
    SECONDARY = '#7000ff'
    SUCCESS = '#00ff9d'
    DANGER = '#ff2e63'
    WARNING = '#ffaa00'
    INFO = '#aa00ff'
    TEXT_MAIN = '#e1e1e1'
    TEXT_DIM = '#71767b'
    BORDER = '#1f242d'

# ==================== فئة الزر التفاعلي ====================
class HoverButton(tk.Button):
    def __init__(self, master, hover_bg, **kw):
        tk.Button.__init__(self, master=master, **kw)
        self.default_bg = self['bg']
        self.hover_bg = hover_bg
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
    
    def on_enter(self, e):
        self['bg'] = self.hover_bg
    
    def on_leave(self, e):
        self['bg'] = self.default_bg

# ==================== نافذة كلمة المرور البسيطة ====================
class PasswordDialog(simpledialog.Dialog):
    """نافذة كلمة المرور — مُوسَّطة دائماً"""
    def __init__(self, parent, title, message, username):
        self.username       = username
        self.custom_message = message
        super().__init__(parent, title)

    def body(self, master):
        self.configure(bg=Colors.BG_CARD)
        master.configure(bg=Colors.BG_CARD)
        tk.Label(master, text=self.custom_message, font=("Segoe UI", 10),
                 fg=Colors.TEXT_MAIN, bg=Colors.BG_CARD,
                 justify=tk.CENTER).pack(pady=15)
        self.entry = tk.Entry(master, show="●", font=("Segoe UI", 11),
                              bg=Colors.BG_DARK, fg=Colors.TEXT_MAIN,
                              insertbackground='white', width=30)
        self.entry.pack(pady=10)
        self.entry.focus_set()
        self.after(10, self._center_self)
        return self.entry

    def _center_self(self):
        self.update_idletasks()
        w  = self.winfo_width()  or 350
        h  = self.winfo_height() or 200
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"+{(sw - w) // 2}+{(sh - h) // 2}")

    def apply(self):
        self.result = self.entry.get()

# ==================== نظام الإشعارات باستخدام Plyer ====================
def show_notification(title, message):
    """إرسال إشعار باستخدام plyer مع fallback"""
    try:
        # محاولة استخدام plyer
        from plyer import notification
        notification.notify(
            title=title,
            message=message,
            app_name="USB Shield Pro",
            timeout=3
        )
        print(f"🔔 Notification sent: {title}")
    except Exception as e:
        print(f"⚠️ Plyer notification failed: {e}")
        # Fallback: استخدم messagebox غير مزعج
        try:
            root_temp = tk.Tk()
            root_temp.withdraw()
            root_temp.attributes('-topmost', True)
            root_temp.after(1000, root_temp.destroy)
            messagebox.showinfo(title, message, parent=root_temp)
        except:
            print(f"📝 {title}: {message}")



# ==================== المراقبة المستمرة ====================
# class USBMonitor:
#     """فئة للمراقبة المستمرة لأجهزة USB - مُصلحة"""
#     def __init__(self, callback=None):
#         self.running = False
#         self.monitor_thread = None
#         self.callback = callback
#         self.known_devices = set()

#     def start(self):
#         """بدء المراقبة في خلفية التطبيق"""
#         if self.running:
#             return
        
#         self.running = True
#         self.monitor_thread = threading.Thread(target=self._monitor_loop, daemon=True)
#         self.monitor_thread.start()
#         print("✅ USB Monitor started")

#     def stop(self):
#         """إيقاف المراقبة"""
#         self.running = False
#         if self.monitor_thread:
#             self.monitor_thread.join()
#         print("⏹️ USB Monitor stopped")

#     def _monitor_loop(self):
#         """حلقة المراقبة المستمرة - مع تهيئة COM"""
#         # ⚠️  هذا السطر هو الحل! تهيئة COM داخل الـ Thread
#         pythoncom.CoInitialize()
        
#         try:
#             while self.running:
#                 c = wmi.WMI()
#                 current_devices = set()
                
#                 for disk in c.Win32_DiskDrive(InterfaceType="USB"):
#                     serial = disk.SerialNumber.strip() if disk.SerialNumber else None
#                     if serial:
#                         current_devices.add(serial)
                
#                 # كشف الأجهزة الجديدة
#                 new_devices = current_devices - self.known_devices
#                 removed_devices = self.known_devices - current_devices
                
#                 if new_devices and self.callback:
#                     for serial in new_devices:
#                         self.callback("new", serial)
                
#                 if removed_devices and self.callback:
#                     for serial in removed_devices:
#                         self.callback("removed", serial)
                
#                 self.known_devices = current_devices
#                 time.sleep(2)  # فحص كل ثانيتين
                
#         except Exception as e:
#             print(f"❌ Monitor error: {e}")
#         finally:
#             # تنظيف COM عند الخروج
#             pythoncom.CoUninitialize()

# ==================== مدير الأجهزة الموحد ====================
class DeviceListManager:
    """فئة موحدة لإدارة عمليات النقل بين القوائم"""
    @staticmethod
    def move_device(fingerprint, from_list, to_list, user, reason=""):
        try:
            conn = get_connection()
            cursor = conn.cursor()

            # ── قراءة بيانات الجهاز من القائمة المصدر ──────────────────────────
            source_query = {
                'whitelist':    'SELECT * FROM whitelist    WHERE fingerprint = ?',
                'blacklist':    'SELECT * FROM blacklist    WHERE fingerprint = ?',
                'auto_blocked': 'SELECT * FROM auto_blocked WHERE fingerprint = ?',
            }
            if from_list not in source_query:
                conn.close(); return False, "Invalid source list"

            cursor.execute(source_query[from_list], (fingerprint,))
            row = cursor.fetchone()
            if not row:
                conn.close(); return False, "Device not found"

            columns     = [desc[0] for desc in cursor.description]
            device_data = dict(zip(columns, row))
            now_local   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # ── الإضافة للقائمة الهدف ────────────────────────────────────────
            success      = False
            pnp_for_unblock = (device_data.get('pnp_device_id') or
                               device_data.get('pnp_id') or '')

            if to_list == 'whitelist':
                whitelist_data = {
                    'fingerprint': fingerprint,
                    'basic_info': {
                        'model':         device_data.get('model', 'Unknown'),
                        'serial_number': device_data.get('serial_number', 'N/A'),
                        'size_gb':       device_data.get('size_gb', 0),
                    },
                    'hardware_info': {
                        'vendor_id':  device_data.get('vid', 'N/A'),
                        'product_id': device_data.get('pid', 'N/A'),
                    }
                }
                success = add_to_whitelist(
                    whitelist_data, added_by=user,
                    notes=f"Moved from {from_list}: {reason}"
                )

            elif to_list == 'blacklist':
                blacklist_data = {
                    'fingerprint': fingerprint,
                    'basic_info': {
                        'model':         device_data.get('model', 'Unknown'),
                        'serial_number': device_data.get('serial_number', 'N/A'),
                        'size_gb':       device_data.get('size_gb', 0),
                    },
                    'hardware_info': {
                        'vendor_id':  device_data.get('vid', 'N/A'),
                        'product_id': device_data.get('pid', 'N/A'),
                    }
                }
                success = add_to_blacklist(
                    blacklist_data, blocked_by=user,
                    block_reason=f"Moved from {from_list}: {reason}"
                )

            elif to_list == 'auto_blocked':
                cursor.execute('''
                    INSERT OR REPLACE INTO auto_blocked
                    (fingerprint, model, serial_number, vid, pid, size_gb,
                     pnp_device_id, block_reason, blocked_by, first_seen, last_seen)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'SYSTEM', ?, ?)
                ''', (
                    fingerprint,
                    device_data.get('model'),
                    device_data.get('serial_number'),
                    device_data.get('vid'),
                    device_data.get('pid'),
                    device_data.get('size_gb'),
                    pnp_for_unblock,
                    f"Moved from {from_list}: {reason}",
                    now_local, now_local
                ))
                conn.commit()
                success = True
            else:
                conn.close(); return False, "Invalid target list"

            if not success:
                conn.close(); return False, "Failed to add to target list"

            # ── حذف من كل القوائم الأخرى (ليس المصدر فقط) ──────────────────
            # يضمن عدم وجود الجهاز في أكثر من قائمة واحدة في أي وقت
            for tbl in ('whitelist', 'blacklist', 'auto_blocked'):
                if tbl != to_list:
                    cursor.execute(f'DELETE FROM {tbl} WHERE fingerprint = ?', (fingerprint,))
            conn.commit()
            conn.close()

            log_event(
                event_type=f"DEVICE_MOVED_{from_list.upper()}_TO_{to_list.upper()}",
                device_fingerprint=fingerprint,
                device_model=device_data.get('model', 'Unknown'),
                result="Success", user=user,
                details=f"Moved from {from_list} to {to_list}. Reason: {reason}"
            )

            # ── تطبيق الحظر / رفعه فعلياً بعد تحديث قاعدة البيانات ──────────
            if to_list == 'whitelist':
                serial_to_find = (device_data.get('serial_number') or
                                  device_data.get('serial') or '')

                # نبحث عن pnp_id بثلاث طرق بالترتيب:
                # 1. من قاعدة البيانات مباشرة (auto_blocked أو blacklist مع pnp_device_id)
                # 2. Win32_DiskDrive (جهاز نشط)
                # 3. Win32_PnPEntity (جهاز معطّل في Device Manager)
                if not pnp_for_unblock and serial_to_find and serial_to_find != 'N/A':
                    try:
                        import pythoncom as _pycom
                        _pycom.CoInitialize()
                        c = wmi.WMI()

                        # محاولة 1: الأجهزة النشطة
                        for disk in c.Win32_DiskDrive(InterfaceType="USB"):
                            if (disk.SerialNumber or "").strip() == serial_to_find:
                                pnp_for_unblock = getattr(disk, 'PNPDeviceID', None)
                                print(f"  Found active device: {pnp_for_unblock}")
                                break

                        # محاولة 2: الأجهزة المعطّلة — نبحث عن USB parent بالـ serial
                        if not pnp_for_unblock:
                            serial_upper = serial_to_find.upper()
                            for pnp in c.Win32_PnPEntity():
                                pid_str = (getattr(pnp, 'PNPDeviceID', '') or '').upper()
                                if serial_upper in pid_str and pid_str.startswith('USB\\VID_'):
                                    pnp_for_unblock = getattr(pnp, 'PNPDeviceID', None)
                                    print(f"  Found disabled USB parent: {pnp_for_unblock}")
                                    break

                    except Exception as e:
                        print(f"⚠️ WMI search: {e}")

                # unblock_and_reenable تُمكّن كل الأجهزة المرتبطة بالـ serial
                # سواء أعطيناها USB parent أو USBSTOR أو حتى None (ستعمل بالـ serial فقط)
                if pnp_for_unblock or serial_to_find:
                    target = pnp_for_unblock or f"SERIAL:{serial_to_find}"
                    print(f"🔓 Auto-enabling via: {target}")
                    USBBlocker.unblock_and_reenable(
                        pnp_for_unblock if pnp_for_unblock else serial_to_find
                    )
                    return True, "WHITELIST_AUTO_ENABLED"
                else:
                    return True, "WHITELIST_REPLUG_REQUIRED"

            elif to_list == 'blacklist':
                # تطبيق الحظر الفعلي فوراً على الجهاز المتصل حالياً
                if pnp_for_unblock:
                    print(f"⛔ Applying real-time block: {pnp_for_unblock}")
                    from usb_blocker import add_to_deny_list, disable_via_pnputil
                    add_to_deny_list(pnp_for_unblock)
                    disable_via_pnputil(pnp_for_unblock)
                    return True, "BLACKLIST_BLOCKED"
                else:
                    # الجهاز ليس متصلاً، سيُحظر تلقائياً عند التوصيل
                    return True, "BLACKLIST_WILL_BLOCK_ON_CONNECT"

            return True, "Success"

        except Exception as e:
            print(f"❌ Error moving device: {e}")
            return False, str(e)

    @staticmethod
    def get_auto_blocked_devices():
        """جلب جميع الأجهزة في قائمة الحظر التلقائي"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM auto_blocked ORDER BY last_seen DESC')
            devices = []
            # استخدام أسماء الأعمدة بدلاً من المواضع لضمان التوافق
            cols = [d[0] for d in cursor.description]
            for row in cursor.fetchall():
                device = dict(zip(cols, row))
                devices.append(device)
            conn.close()
            return devices
        except Exception as e:
            print(f"Error getting auto-blocked devices: {e}")
            return []

    @staticmethod
    def add_to_auto_blocked(device_info, reason="Auto-detected unrecognized device"):
        """إضافة جهاز جديد للقائمة المؤقتة عند اكتشافه"""
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # 1. الحصول على الوقت الحالي بدقة من بايثون
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # 2. التحقق مما إذا كان الجهاز موجوداً بالفعل
            cursor.execute("SELECT first_seen FROM auto_blocked WHERE fingerprint = ?", (device_info['fingerprint'],))
            existing = cursor.fetchone()
            
            if existing:
                # تحديث last_seen فقط إذا كان الجهاز موجوداً سابقاً
                cursor.execute('''
                    UPDATE auto_blocked SET last_seen = ? WHERE fingerprint = ?
                ''', (now, device_info['fingerprint']))
            else:
                # إدراج سجل جديد بالكامل
                cursor.execute('''
                    INSERT INTO auto_blocked 
                    (fingerprint, model, serial_number, vid, pid, size_gb, block_reason, blocked_by, first_seen, last_seen)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    device_info['fingerprint'],
                    device_info['model'],
                    device_info['serial_number'],
                    device_info.get('vid', 'N/A'),
                    device_info.get('pid', 'N/A'), 
                    device_info['size_gb'],
                    reason,
                    'SYSTEM',
                    now,  # وقت الظهور الأول
                    now   # وقت آخر ظهور
                ))
            
            conn.commit()
            conn.close()
            
            # تسجيل الحدث
            log_event(
                event_type="DEVICE_AUTO_BLOCKED",
                device_fingerprint=device_info['fingerprint'],
                device_model=device_info['model'],
                result="Success",
                user="SYSTEM",
                details=f"Unrecognized device auto-blocked: {reason}"
            )
            return True
            
        except Exception as e:
            print(f"Error adding to auto_blocked: {e}")
            return False

# ==================== دوال مساعدة ====================


def extract_vid_pid_smart(pnp_id, serial):
    """استخراج ذكي لـ VID و PID مع دعم مسارات USBSTOR"""
    vid, pid = "N/A", "N/A"
    if not pnp_id: return vid, pid

    # 1️⃣ محاولة الاستخراج المباشر من صيغة USB القياسية
    direct_match = re.search(r'USB\\VID_([0-9A-Fa-f]{4})&PID_([0-9A-Fa-f]{4})', pnp_id)
    if direct_match:
        return direct_match.group(1), direct_match.group(2)

    # 2️⃣ إذا كانت الصيغة USBSTOR، نبحث عن الجهاز الأب في نظام الويندوز
    if pnp_id.startswith("USBSTOR") and serial != "N/A":
        try:
            c = wmi.WMI()
            for pnp in c.Win32_PnPEntity():
                dev_pnp = getattr(pnp, 'PNPDeviceID', '') or ''
                # مطابقة الرقم التسلسلي أو معرف الجهاز
                if serial in dev_pnp and dev_pnp.startswith("USB\\VID_"):
                    vm = re.search(r'VID_([0-9A-Fa-f]{4})', dev_pnp)
                    pm = re.search(r'PID_([0-9A-Fa-f]{4})', dev_pnp)
                    if vm: vid = vm.group(1)
                    if pm: pid = pm.group(1)
                    break
        except Exception:
            pass
    return vid, pid

def generate_fingerprint_from_device(disk):
    try:
        serial = (disk.SerialNumber or "").strip() or "NO_SERIAL"
        size = str(round(int(disk.Size) / (1024 ** 3), 2)) if disk.Size else "0"
        pnp_id = getattr(disk, 'PNPDeviceID', '') or ''
        
        # استخراج ذكي للـ VID/PID داخل دالة البصمة
        vid, pid = "N/A", "N/A"
        vm = re.search(r'VID_([0-9A-Fa-f]{4})', pnp_id)
        pm = re.search(r'PID_([0-9A-Fa-f]{4})', pnp_id)
        if vm: vid = vm.group(1)
        if pm: pid = pm.group(1)
            
        raw = f"{serial}{vid}{pid}_{size}"
        return hashlib.sha256(raw.encode()).hexdigest()
    except:
        return None

def get_connected_usb_devices():
    try:
        c = wmi.WMI()
        disks = c.Win32_DiskDrive(InterfaceType="USB")
        if not disks: return []
        usb_drives = []
        for disk in disks:
            pnp_id = getattr(disk, 'PNPDeviceID', None)
            serial = (disk.SerialNumber or "").strip() or "N/A"
            size_gb = round(int(disk.Size) / (1024 ** 3), 2) if disk.Size else 0
            model = (disk.Model or "Unknown").strip()
            
            vid, pid = extract_vid_pid_smart(pnp_id, serial)
            raw_fp = f"{serial}{vid}{pid}_{size_gb}"
            fingerprint = hashlib.sha256(raw_fp.encode()).hexdigest()
            
            usb_drives.append({
                'model': model, 'device_id': disk.DeviceID, 'pnp_device_id': pnp_id,
                'serial_number': serial, 'size_gb': size_gb,
                'fingerprint': fingerprint, 'vid': vid, 'pid': pid
            })
        return usb_drives
    except Exception as e:
        print(f"❌ Error getting USB devices: {e}")
        return []
        

def get_device_status(fingerprint):
    """تحديد حالة الجهاز بدقة من القوائم الثلاث"""
    if not fingerprint:
        return "Unknown"
    if is_device_whitelisted(fingerprint):
        return "Whitelisted"
    if device_exists_in_blacklist(fingerprint):
        return "Blacklisted"
    return "Blocked"  # موجود في auto_blocked أو غير معروف

# ==================== نافذة تسجيل الدخول ====================
class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("USB Shield - Login")
        self.root.geometry("400x550")
        self.root.configure(bg=Colors.BG_DARK)
        self.create_widgets()
        self.center_window()
    
    def center_window(self):
        self.root.update_idletasks()
        width, height = 400, 550
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        header = tk.Frame(self.root, bg=Colors.BG_DARK)
        header.pack(pady=40)
        tk.Label(header, text="🛡️", font=("Segoe UI Emoji", 50), fg=Colors.PRIMARY, bg=Colors.BG_DARK).pack()
        tk.Label(header, text="USB SHIELD", font=("Segoe UI", 22, "bold"), fg=Colors.TEXT_MAIN, bg=Colors.BG_DARK).pack(pady=10)
        tk.Label(header, text="Sign in To Access USB Control", font=("Segoe UI", 9), fg=Colors.DANGER, bg=Colors.BG_DARK).pack()

        form_frame = tk.Frame(self.root, bg=Colors.BG_DARK)
        form_frame.pack(padx=40, fill=tk.X)

        tk.Label(form_frame, text="USERNAME", font=("Segoe UI", 8, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_DARK).pack(anchor='w', pady=(20, 5))
        self.username_entry = tk.Entry(form_frame, font=("Segoe UI", 12), bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN, insertbackground='white', relief='flat', bd=10)
        self.username_entry.pack(fill=tk.X)

        tk.Label(form_frame, text="PASSWORD", font=("Segoe UI", 8, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_DARK).pack(anchor='w', pady=(15, 5))
        self.password_entry = tk.Entry(form_frame, font=("Segoe UI", 12), show="●", bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN, insertbackground='white', relief='flat', bd=10)
        self.password_entry.pack(fill=tk.X)

        self.login_btn = HoverButton(self.root, text="AUTHENTICATE", font=("Segoe UI", 11, "bold"),
                                    bg=Colors.PRIMARY, fg=Colors.BG_DARK, hover_bg='#00b8e6',
                                    relief='flat', bd=0, cursor='hand2', command=self.do_login)
        self.login_btn.pack(pady=40, padx=40, fill=tk.X, ipady=10)

        self.error_label = tk.Label(self.root, text="", font=("Segoe UI", 9), fg=Colors.DANGER, bg=Colors.BG_DARK)
        self.error_label.pack()

        self.username_entry.bind('<Return>', lambda e: self.password_entry.focus())
        self.password_entry.bind('<Return>', lambda e: self.do_login())
        self.username_entry.focus()
    
    def do_login(self):
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        if not username or not password:
            self.error_label.config(text="⚠️ Identification required")
            return
        result = login(username, password)
        if result['success']:
            # ── حفظ اسم المستخدم وتفعيل بدء التشغيل التلقائي ──────────
            # يجب حفظ last_username و auto_start_enabled قبل enable_startup
            # حتى يعمل _is_auto_start_enabled عند إقلاع الويندوز
            try:
                from database_manager import set_setting
                set_setting('last_username', username)
                set_setting('auto_start_enabled', '1')
                print(f"✅ Session saved: user={username}, auto_start=1")
            except Exception as e:
                print(f"⚠️ Could not save session settings: {e}")
            # ── تسجيل/تحديث Startup Registry ────────────────────────────
            # نُسجّل دائماً (وليس فقط إذا لم يكن مسجّلاً) لضمان صحة المسار
            enable_startup()
            # ── فتح النافذة الرئيسية ─────────────────────────────────────
            self.root.destroy()
            new_root = tk.Tk()
            app = MainWindow(new_root, username)
            new_root.mainloop()
        else:
            self.error_label.config(text=f"❌ Access Denied: {result['message']}")

# ==================== النافذة الرئيسية ====================
class MainWindow:
    def __init__(self, root, username, start_hidden=False):
        self.username     = username
        self.root         = root
        self.tray_icon    = None
        self._tray_thread = None
        self.start_hidden = start_hidden
        self.root.title(f"USB Shield - {username}")
        self.root.configure(bg=Colors.BG_DARK)
        self.current_view = 'dashboard'

        w, h = 1300, 800
        self.root.update_idletasks()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.geometry(f"{w}x{h}+{(sw - w) // 2}+{(sh - h) // 2}")

        self.usb_monitor = USBMonitor(on_device_change=self.on_usb_change, tk_root=self.root)
        self.usb_monitor.start()

        autoplay_result = USBBlocker.disable_autoplay_if_enabled()

        self.setup_styles()
        self.create_layout()
        self.show_dashboard()

        self._start_tray()

        if start_hidden:
            # تشغيل صامت من Startup — لا نافذة، لا إشعار
            self.root.withdraw()
        else:
            self.root.after(800, lambda: self._notify_autoplay_status(autoplay_result))

        self.root.protocol("WM_DELETE_WINDOW", self._hide_to_tray)

        # ── ضمان تسجيل Startup إذا لم يكن مسجلاً (حالة headless) ─────────
        if not is_startup_enabled():
            enable_startup()
            print("✅ Registered in Windows startup")

    # ==================== System Tray ====================

    def _start_tray(self):
        """يُشغِّل أيقونة الـ tray — زرين فقط: Open و Exit"""
        if not TRAY_AVAILABLE:
            print("⚠️ pystray/PIL not installed — run: pip install pystray Pillow")
            return
        try:
            icon_img = _make_tray_icon(64)
            if not isinstance(icon_img, Image.Image):
                raise ValueError("Icon generation failed")
            menu = pystray.Menu(
                pystray.MenuItem("🛡️ USB Shield", None, enabled=False),
                pystray.Menu.SEPARATOR,
                pystray.MenuItem("Open",  self._tray_open),
                pystray.MenuItem("Exit",  self._tray_exit),
            )
            self.tray_icon = pystray.Icon(
                name  = "USBShield",
                icon  = icon_img,
                title = "USB Shield — Protection Active",
                menu  = menu,
            )
            self._tray_thread = threading.Thread(
                target=self.tray_icon.run,
                daemon=True, name="TrayThread"
            )
            self._tray_thread.start()
            print("✅ System Tray icon started successfully")
        except Exception as e:
            print(f"❌ Tray failed to start: {e}")
            import traceback; traceback.print_exc()

    def _hide_to_tray(self):
        self.root.withdraw()

    def _tray_open(self, icon=None, item=None):
        """فتح النافذة — يطلب كلمة مرور دائماً"""
        self.root.after(0, self._ask_password_then_open)

    def _ask_password_then_open(self):
        self._show_auth_dialog(
            title="Open USB Shield",
            subtitle="Enter password to access the dashboard",
            on_success=self._restore_window
        )

    def _restore_window(self):
        self.root.deiconify()
        self.root.lift()
        self.root.focus_force()

    def _tray_exit(self, icon=None, item=None):
        """إغلاق البرنامج — يطلب كلمة مرور"""
        self.root.after(0, self._ask_password_then_exit)

    def _ask_password_then_exit(self):
        self._show_auth_dialog(
            title="Exit USB Shield",
            subtitle="USB protection will stop. Enter password to confirm.",
            on_success=self._do_full_exit,
            danger=True
        )

    def _show_auth_dialog(self, title, subtitle, on_success, danger=False):
        """نافذة كلمة مرور موحّدة تُستخدم لـ Open و Exit"""
        win = tk.Toplevel()
        win.title(f"USB Shield — {title}")
        win.configure(bg=Colors.BG_DARK)
        win.resizable(False, False)
        win.grab_set()
        win.focus_force()
        w, h = 360, 290
        win.update_idletasks()
        sw, sh = win.winfo_screenwidth(), win.winfo_screenheight()
        win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

        icon_txt = "⚠️" if danger else "🛡️"
        icon_color = Colors.WARNING if danger else Colors.PRIMARY
        tk.Label(win, text=icon_txt, font=("Segoe UI Emoji", 28),
                 fg=icon_color, bg=Colors.BG_DARK).pack(pady=(20, 4))
        tk.Label(win, text=title, font=("Segoe UI", 13, "bold"),
                 fg=Colors.TEXT_MAIN, bg=Colors.BG_DARK).pack()
        tk.Label(win, text=subtitle, font=("Segoe UI", 9),
                 fg=Colors.TEXT_DIM, bg=Colors.BG_DARK, wraplength=300).pack(pady=(3, 12))

        pwd_var = tk.StringVar()
        entry = tk.Entry(win, textvariable=pwd_var, show="●",
                         font=("Segoe UI", 12), bg=Colors.BG_CARD,
                         fg=Colors.TEXT_MAIN, insertbackground='white',
                         relief='flat', bd=8)
        entry.pack(padx=30, fill=tk.X)
        entry.focus()

        err_lbl = tk.Label(win, text="", font=("Segoe UI", 9),
                           fg=Colors.DANGER, bg=Colors.BG_DARK)
        err_lbl.pack(pady=3)

        def _verify():
            from auth_manager import verify_password
            if verify_password(self.username, pwd_var.get().strip()):
                win.destroy()
                on_success()
            else:
                err_lbl.config(text="❌ Incorrect password")
                pwd_var.set(""); entry.focus()

        btn_color = Colors.DANGER if danger else Colors.PRIMARY
        btn_hover = '#c0392b' if danger else '#00b8e6'
        HoverButton(win, text="CONFIRM", font=("Segoe UI", 10, "bold"),
                    bg=btn_color, fg='white' if danger else Colors.BG_DARK,
                    hover_bg=btn_hover, relief='flat', bd=0, cursor='hand2',
                    command=_verify).pack(padx=30, fill=tk.X, ipady=8)
        entry.bind('<Return>', lambda e: _verify())
        entry.bind('<Escape>', lambda e: win.destroy())

    def _do_full_exit(self):
        """
        إغلاق حقيقي كامل من قائمة الـ Tray — يطلب كلمة مرور.
        يحذف من Startup ويمسح علامة التشغيل التلقائي.
        المرة القادمة سيطلب Login من جديد.
        """
        try:
            from database_manager import set_setting
            set_setting('auto_start_enabled', '0')
        except Exception:
            pass
        disable_startup()
        USBBlocker.restore_autoplay()
        if self.tray_icon:
            try: self.tray_icon.stop()
            except Exception: pass
        self.usb_monitor.stop()
        self.root.quit()
        self.root.destroy()

    def _on_close(self):
        """إغلاق نهائي (يُستدعى من Logout أو Exit)"""
        # ① استعادة AutoPlay لحالته الأصلية قبل الإغلاق
        USBBlocker.restore_autoplay()
        if self.tray_icon:
            try:
                self.tray_icon.stop()
            except Exception:
                pass
        self.usb_monitor.stop()
        self.root.destroy()

    # ==================== إشعارات AutoPlay ====================

    def _notify_autoplay_status(self, result: str):
        """
        يعرض إشعار toast غير مُزعج يُخبر المستخدم بحالة AutoPlay
        فور بدء التشغيل.
          DISABLED_NOW  — كان مُفعَّلاً وتم إيقافه الآن  (تحذير + إشعار نظام)
          ALREADY_OFF   — كان مُطفأً مسبقاً             (رسالة هادئة خضراء)
          FAILED        — فشل الإيقاف                   (تنبيه أحمر + إشعار نظام)
        """
        if result == "DISABLED_NOW":
            icon  = "🛡️"
            title = "AutoPlay Disabled"
            msg   = "AutoPlay was ON — disabled automatically to protect your system."
            color = Colors.WARNING
            show_notification("USB Shield — AutoPlay Disabled",
                              "AutoPlay was ON and has been disabled automatically.")
        elif result == "ALREADY_OFF":
            icon  = "✅"
            title = "AutoPlay Protected"
            msg   = "AutoPlay is already disabled. Your system is protected."
            color = Colors.SUCCESS
        else:  # FAILED
            icon  = "⚠️"
            title = "AutoPlay Warning"
            msg   = "Could not disable AutoPlay. Please disable it manually from Windows Settings."
            color = Colors.DANGER
            show_notification("USB Shield — AutoPlay Warning",
                              "Failed to disable AutoPlay. Please check your Windows Settings.")

        self._show_toast(icon, title, msg, color)

    def _show_toast(self, icon: str, title: str, message: str,
                    color: str, duration_ms: int = 5000):
        """
        إشعار toast يظهر في الزاوية اليمنى السفلى من النافذة الرئيسية
        ويتحرك معها — ويختفي تلقائياً بعد duration_ms.
        """
        toast = tk.Toplevel(self.root)
        toast.overrideredirect(True)
        toast.attributes('-topmost', True)
        toast.configure(bg=Colors.BG_CARD)

        # ── بناء المحتوى ───────────────────────────────────────────────────
        outer = tk.Frame(toast, bg=color, padx=2, pady=2)
        outer.pack(fill=tk.BOTH, expand=True)

        inner = tk.Frame(outer, bg=Colors.BG_CARD, padx=16, pady=12)
        inner.pack(fill=tk.BOTH, expand=True)

        header = tk.Frame(inner, bg=Colors.BG_CARD)
        header.pack(fill=tk.X)

        tk.Label(header, text=icon, font=("Segoe UI Emoji", 16),
                 fg=color, bg=Colors.BG_CARD).pack(side=tk.LEFT, padx=(0, 8))
        tk.Label(header, text=title, font=("Segoe UI", 10, "bold"),
                 fg=color, bg=Colors.BG_CARD).pack(side=tk.LEFT)

        tk.Button(header, text="✕", font=("Segoe UI", 8),
                  fg=Colors.TEXT_DIM, bg=Colors.BG_CARD,
                  relief='flat', bd=0, cursor='hand2',
                  command=toast.destroy).pack(side=tk.RIGHT)

        tk.Label(inner, text=message, font=("Segoe UI", 9),
                 fg=Colors.TEXT_MAIN, bg=Colors.BG_CARD,
                 wraplength=300, justify=tk.LEFT).pack(anchor='w', pady=(6, 0))

        TOAST_W = 340

        def _reposition():
            """يُعيد حساب موضع الـ toast بناءً على موضع النافذة الحالي."""
            if not toast.winfo_exists():
                return
            try:
                self.root.update_idletasks()
                toast.update_idletasks()
                toast_h = toast.winfo_reqheight() + 4
                pos_x = self.root.winfo_x() + self.root.winfo_width()  - TOAST_W - 20
                pos_y = self.root.winfo_y() + self.root.winfo_height() - toast_h - 20
                toast.geometry(f"{TOAST_W}x{toast_h}+{pos_x}+{pos_y}")
            except Exception:
                pass

        # موضع أولي
        _reposition()

        # ── تتبع حركة النافذة الرئيسية ────────────────────────────────────
        _tracking = [True]   # قائمة للتعديل داخل الـ closure

        def _track_loop():
            if not _tracking[0] or not toast.winfo_exists():
                return
            _reposition()
            toast.after(100, _track_loop)   # تحديث كل 100ms

        toast.after(100, _track_loop)

        # ── اختفاء تلقائي ──────────────────────────────────────────────────
        def _fade_out():
            _tracking[0] = False
            try:
                toast.destroy()
            except Exception:
                pass

        toast.after(duration_ms, _fade_out)

    def on_usb_change(self, action, data):
        if action == "new":
            if isinstance(data, dict):
                serial = data.get('serial', 'N/A')
                model  = data.get('model', 'Unknown USB')
                print(f"🔔 New USB device detected: {serial}")
                # بعد 2.5 ثانية — نتحقق هل تم حظره وعرض toast مناسب
                self.root.after(2500, lambda: self._check_and_notify_block(data))
            self.root.after(2000, self.refresh_if_needed)
            self.root.after(2000, self.refresh_auto_blocked_if_needed)
        elif action == "removed":
            print(f"🔔 USB device removed")
            self.root.after(0, self.refresh_if_needed)

    def _check_and_notify_block(self, device_data):
        """يفحص ما إذا تم حظر الجهاز ويعرض toast مناسب مع زر Trust"""
        try:
            import hashlib
            serial = device_data.get('serial', 'N/A')
            vid    = device_data.get('vid', 'N/A')
            pid    = device_data.get('pid', 'N/A')
            size   = device_data.get('size_gb', 0)
            model  = device_data.get('model', 'Unknown')
            fp     = hashlib.sha256(f"{serial}{vid}{pid}_{size}".encode()).hexdigest()
            status = get_device_status(fp)

            if status == "Whitelisted":
                self._show_toast("✅", "Device Allowed", f"{model[:28]} — Authorized", Colors.SUCCESS, duration_ms=3000)
            elif status == "Blacklisted":
                self._show_block_toast(model, fp, "Blacklisted")
            else:
                self._show_block_toast(model, fp, "Blocked")
        except Exception as e:
            print(f"⚠️ Notify error: {e}")

    def _show_block_toast(self, model, fingerprint, status):
        """Toast احترافي عند الحظر مع زر Trust مباشر"""
        try:
            toast = tk.Toplevel(self.root)
            toast.overrideredirect(True)
            toast.configure(bg=Colors.BG_CARD)
            toast.attributes('-topmost', True)

            # توضيع في أسفل اليمين
            toast.update_idletasks()
            sw = toast.winfo_screenwidth()
            sh = toast.winfo_screenheight()
            w, h = 300, 110
            toast.geometry(f"{w}x{h}+{sw-w-20}+{sh-h-60}")

            # حدود
            border = tk.Frame(toast, bg=Colors.DANGER, padx=1, pady=1)
            border.pack(fill=tk.BOTH, expand=True)
            inner = tk.Frame(border, bg=Colors.BG_CARD, padx=14, pady=10)
            inner.pack(fill=tk.BOTH, expand=True)

            # السطر الأول: أيقونة + عنوان
            top_row = tk.Frame(inner, bg=Colors.BG_CARD)
            top_row.pack(fill=tk.X)
            tk.Label(top_row, text="⛔", font=("Segoe UI Emoji", 12),
                     fg=Colors.DANGER, bg=Colors.BG_CARD).pack(side=tk.LEFT)
            tk.Label(top_row, text="  USB Blocked",
                     font=("Segoe UI", 10, "bold"),
                     fg=Colors.TEXT_MAIN, bg=Colors.BG_CARD).pack(side=tk.LEFT)

            # اسم الجهاز
            tk.Label(inner, text=model[:34],
                     font=("Segoe UI", 8), fg=Colors.TEXT_DIM,
                     bg=Colors.BG_CARD, anchor='w').pack(fill=tk.X, pady=(2, 6))

            # الأزرار
            btn_row = tk.Frame(inner, bg=Colors.BG_CARD)
            btn_row.pack(fill=tk.X)

            def _trust():
                toast.destroy()
                # نقل للبيضاء مباشرة بعد التحقق من كلمة المرور
                if self.verify_user_password(f"Trust: {model[:25]}"):
                    result, msg = DeviceListManager.move_device(
                        fingerprint, 'auto_blocked', 'whitelist',
                        self.username, reason="Trusted via notification"
                    )
                    if result:
                        self._show_toast("✅", "Device Trusted", f"{model[:28]} added to whitelist", Colors.SUCCESS)
                        self.refresh_if_needed()

            HoverButton(btn_row, text="✅ Trust",
                        font=("Segoe UI", 8, "bold"),
                        bg=Colors.SUCCESS, fg=Colors.BG_DARK, hover_bg='#00cc7a',
                        relief='flat', bd=0, padx=8, pady=3,
                        command=_trust).pack(side=tk.LEFT)
            HoverButton(btn_row, text="Dismiss",
                        font=("Segoe UI", 8),
                        bg=Colors.BG_CARD, fg=Colors.TEXT_DIM, hover_bg=Colors.SIDEBAR,
                        relief='flat', bd=0, padx=8, pady=3,
                        command=toast.destroy).pack(side=tk.LEFT, padx=(6, 0))

            toast.after(8000, lambda: toast.destroy() if toast.winfo_exists() else None)
        except Exception as e:
            print(f"⚠️ Block toast error: {e}")
    
    def refresh_auto_blocked_if_needed(self):
        """تحديث قائمة Auto-Blocked إذا كانت مفتوحة"""
        if hasattr(self, 'current_view') and self.current_view == 'auto_blocked':
            self.show_auto_blocked()
    
    def refresh_if_needed(self):
        """تحديث الواجهة إذا كانت في صفحة Connected USB"""
        if hasattr(self, 'current_view') and self.current_view == 'connected_usb':
            self.show_connected_usb()
    
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('default')
        style.configure("Treeview", bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN, fieldbackground=Colors.BG_CARD, borderwidth=0, rowheight=35)
        style.map("Treeview", background=[('selected', Colors.PRIMARY)], foreground=[('selected', Colors.BG_DARK)])

    def create_layout(self):
        self.sidebar = tk.Frame(self.root, bg=Colors.SIDEBAR, width=260)
        self.sidebar.pack(side=tk.RIGHT, fill=tk.Y)
        self.sidebar.pack_propagate(False)
        self.main_container = tk.Frame(self.root, bg=Colors.BG_DARK)
        self.main_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.create_sidebar_content()

    def create_sidebar_content(self):
        profile = tk.Frame(self.sidebar, bg=Colors.SIDEBAR, pady=30)
        profile.pack(fill=tk.X)
        tk.Label(profile, text="👨‍💻", font=("Segoe UI", 40), fg=Colors.PRIMARY, bg=Colors.SIDEBAR).pack()
        tk.Label(profile, text=self.username.upper(), font=("Segoe UI", 12, "bold"), fg=Colors.TEXT_MAIN, bg=Colors.SIDEBAR).pack(pady=5)
        tk.Label(profile, text="System Administrator", font=("Segoe UI", 8), fg=Colors.TEXT_DIM, bg=Colors.SIDEBAR).pack()
        tk.Frame(self.sidebar, bg=Colors.BORDER, height=1).pack(fill=tk.X, padx=20, pady=10)

        menu_items = [
            ("📊  Dashboard", self.show_dashboard),
            ("🔌  Connected USB", self.show_connected_usb),
            ("📋  Whitelist", self.show_whitelist),
            ("🚫  Blacklist", self.show_blacklist),
            ("⏸️  Auto-Blocked", self.show_auto_blocked),
            ("📜  Audit Logs", self.show_audit_logs),
            ("⚙️  Settings", self.show_settings),
        ]
        for text, cmd in menu_items:
            btn = HoverButton(self.sidebar, text=text, font=("Segoe UI", 10),
                            bg=Colors.SIDEBAR, fg=Colors.TEXT_MAIN, hover_bg=Colors.BORDER,
                            relief='flat', bd=0, padx=30, anchor='w', height=2, cursor='hand2', command=cmd)
            btn.pack(fill=tk.X, pady=2)

        logout_btn = HoverButton(self.sidebar, text="🚪  Logout", font=("Segoe UI", 10, "bold"),
                                bg=Colors.SIDEBAR, fg=Colors.DANGER, hover_bg='#2a0a11',
                                relief='flat', bd=0, padx=30, anchor='w', height=2, cursor='hand2', command=self.logout)
        logout_btn.pack(side=tk.BOTTOM, fill=tk.X, pady=20)

    def clear_view(self):
        for widget in self.main_container.winfo_children():
            widget.destroy()


    def _center_window(self, window, width, height):
        """توسيط أي نافذة منبثقة في منتصف الشاشة بدقة"""
        window.update_idletasks()
        sw = window.winfo_screenwidth()
        sh = window.winfo_screenheight()
        window.geometry(f"{width}x{height}+{(sw - width) // 2}+{(sh - height) // 2}")

    def verify_user_password(self, action_name):
        """التحقق من كلمة المرور قبل العمليات الحساسة"""
        dialog = PasswordDialog(self.root, "🔐 Security Check", 
                               f"Enter your password to:\n{action_name}", 
                               self.username)
        if dialog.result is None:
            return False
        if verify_password(self.username, dialog.result):
            return True
        else: 
            messagebox.showerror("Access Denied", "❌ Incorrect password!")
            log_event(
                event_type="FAILED_AUTH_ATTEMPT",
                device_fingerprint="",
                device_model="",
                result="Failed",
                user=self.username,
                details=f"Failed password verification for: {action_name}"
            )
            return False

    def create_card(self, parent, title, value, color, subtitle):
        card = tk.Frame(parent, bg=Colors.BG_CARD, highlightbackground=Colors.BORDER, highlightthickness=1)
        tk.Label(card, text=title, font=("Segoe UI", 10, "bold"), fg=Colors.TEXT_DIM, bg=Colors.BG_CARD).pack(pady=(15, 0))
        tk.Label(card, text=value, font=("Segoe UI", 35, "bold"), fg=color, bg=Colors.BG_CARD).pack()
        tk.Label(card, text=subtitle, font=("Segoe UI", 8), fg=Colors.TEXT_DIM, bg=Colors.BG_CARD).pack(pady=(0, 15))
        return card



    def export_audit_logs(self):
        """
        الدالة الرئيسية للتصدير: تطلب كلمة المرور ثم تفتح نافذة الحفظ الموحدة
        """
        # 1. التحقق من كلمة المرور
        if not self.verify_user_password("Export Audit Logs"):
            return

        # 2. فتح نافذة حفظ الملفات الموحدة (مثل Word)
        from tkinter import filedialog
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[
                ("Excel Workbook", "*.xlsx"),
                ("Comma Separated Values", "*.csv"),
                ("PDF Report", "*.pdf"),
                ("Plain Text", "*.txt")
            ],
            initialfile=f"AuditLogs_{datetime.now().strftime('%Y%m%d_%H%M')}",
            title="Export Audit Logs"
        )

        if not file_path:
            return

        # 3. توجيه التصدير حسب الامتداد المختار
        try:
            if file_path.endswith('.xlsx'):
                self._export_to_excel(file_path)
            elif file_path.endswith('.csv'):
                self._export_to_csv(file_path)
            elif file_path.endswith('.pdf'):
                self._export_to_pdf(file_path)
            else:
                self._export_to_txt(file_path)
                
            show_notification("Export Success", f"Logs exported successfully!")
            messagebox.showinfo("Success", f"Audit logs exported successfully!\n\nSaved to:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export:\n{str(e)}")




    def _export_to_csv(self, file_path):
        """تصدير بصيغة CSV (متوافق مع Excel)"""
        import csv
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM audit_log ORDER BY timestamp DESC')
        rows = cursor.fetchall()
        conn.close()

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            # كتابة العناوين
            writer.writerow(['Timestamp', 'Event Type', 'User', 'Result', 'Device', 'Details'])
            for row in rows:
                writer.writerow([row['timestamp'], row['event_type'], row['user'], 
                                row['result'], row['device_model'], row['details']])

    def _export_to_txt(self, file_path):
        """تصدير بصيغة نصية منظمة"""
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM audit_log ORDER BY timestamp DESC')
        rows = cursor.fetchall()
        conn.close()

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("="*80 + "\n")
            f.write("USB SECURITY MANAGER - AUDIT LOGS REPORT\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("="*80 + "\n\n")
            
            for row in rows:
                f.write(f"[{row['timestamp']}] {row['event_type']}\n")
                f.write(f"   User: {row['user']} | Result: {row['result']}\n")
                f.write(f"   Device: {row['device_model']}\n")
                f.write(f"   Details: {row['details']}\n")
                f.write("-"*80 + "\n")

    def _export_to_excel(self, file_path):
        """تصدير بصيغة Excel الاحترافية"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            messagebox.showerror("Missing Library", "Please install openpyxl:\npip install openpyxl")
            return

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM audit_log ORDER BY timestamp DESC')
        rows = cursor.fetchall()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        # تنسيقات
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="00d1ff", end_color="00d1ff", fill_type="solid")
        
        # إضافة العناوين
        headers = ['Timestamp', 'Event Type', 'User', 'Result', 'Device', 'Details']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # إضافة البيانات
        for r_idx, row in enumerate(rows, 2):
            ws.cell(row=r_idx, column=1, value=row['timestamp'])
            ws.cell(row=r_idx, column=2, value=row['event_type'])
            ws.cell(row=r_idx, column=3, value=row['user'])
            ws.cell(row=r_idx, column=4, value=row['result'])
            ws.cell(row=r_idx, column=5, value=row['device_model'])
            ws.cell(row=r_idx, column=6, value=row['details'])
            
            # تلوين النتيجة
            result_cell = ws.cell(row=r_idx, column=4)
            if row['result'] == 'Success':
                result_cell.font = Font(color="00ff9d", bold=True)
            else:
                result_cell.font = Font(color="ff2e63", bold=True)

        # ضبط الأعمدة
        for col in range(1, 7):
            ws.column_dimensions[chr(64 + col)].width = 25
            
        wb.save(file_path)

    def _export_to_pdf(self, file_path):
        """تصدير بصيغة PDF رسمية"""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import Paragraph
        except ImportError:
            messagebox.showerror("Missing Library", "Please install reportlab:\npip install reportlab")
            return

        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM audit_log ORDER BY timestamp DESC')
        rows = cursor.fetchall()
        conn.close()

        doc = SimpleDocTemplate(file_path, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', fontSize=16, alignment=1) # Center
        elements.append(Paragraph("USB Security Manager - Audit Log Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        
        # تجهيز جدول البيانات
        data = [['Time', 'Event', 'User', 'Result', 'Device', 'Details']]
        for row in rows:
            data.append([row['timestamp'][:16], row['event_type'][:25], row['user'], 
                        row['result'], row['device_model'][:20], row['details'][:40]])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#ff000d")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        elements.append(table)
        doc.build(elements)


    def clear_audit_logs(self):
        """مسح جميع سجلات الأحداث"""
        # التحقق من كلمة المرور أولاً
        if not self.verify_user_password("Clear All Audit Logs"):
            return
        
        # تأكيد المسح
        confirm = messagebox.askyesno(
            "⚠️ Confirm Clear",
            "Are you sure you want to delete ALL audit logs?\n\n"
            "This action cannot be undone!\n\n"
            "💡 Tip: Export logs first to keep a backup.",
            icon='warning'
        )
        
        if not confirm:
            return
        
        try:
            conn = get_connection()
            cursor = conn.cursor()
            
            # الحصول على عدد السجلات قبل المسح
            cursor.execute('SELECT COUNT(*) FROM audit_log')
            count = cursor.fetchone()[0]
            
            # مسح جميع السجلات
            cursor.execute('DELETE FROM audit_log')
            conn.commit()
            conn.close()
            
            show_notification("Logs Cleared", f"{count} log entries deleted")
            messagebox.showinfo("Success", f"All audit logs cleared successfully!\n\nDeleted {count} entries.")
            
            # تسجيل عملية المسح
            log_event(
                event_type="AUDIT_LOGS_CLEARED",
                result="Success",
                user=self.username,
                details=f"Cleared {count} log entries"
            )
            
            # تحديث العرض
            self.show_audit_logs()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to clear logs:\n{e}")

    def show_dashboard(self):
        self.current_view = 'dashboard'
        self.clear_view()
        header = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40, pady=30)
        header.pack(fill=tk.X)
        tk.Label(header, text="System Overview", font=("Segoe UI", 24, "bold"), fg=Colors.TEXT_MAIN, bg=Colors.BG_DARK).pack(side=tk.LEFT)
        tk.Label(header, text=datetime.now().strftime("%A, %d %B %Y"), font=("Segoe UI", 10), fg=Colors.TEXT_DIM, bg=Colors.BG_DARK).pack(side=tk.RIGHT, pady=10)

        grid_frame = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=30)
        grid_frame.pack(fill=tk.BOTH, expand=True)

        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM whitelist'); whitelist_count = cursor.fetchone()[0]
            cursor.execute('SELECT COUNT(*) FROM blacklist'); blacklist_count = cursor.fetchone()[0]
            cursor.execute('SELECT COUNT(*) FROM auto_blocked'); auto_count = cursor.fetchone()[0]
            cursor.execute('SELECT COUNT(*) FROM audit_log'); audit_count = cursor.fetchone()[0]
            conn.close()
            connected_devices = get_connected_usb_devices()
            blocked_count = sum(1 for d in connected_devices if get_device_status(d['fingerprint']) == "Blacklisted")
        except:
            whitelist_count = blacklist_count = auto_count = audit_count = blocked_count = 0

        usb_status = USBBlocker.get_usb_storage_status()
        status_color = Colors.DANGER if usb_status == "Disabled" else Colors.SUCCESS
        status_text = "🔒 BLOCKED" if usb_status == "Disabled" else "🔓 ENABLED"

        # حالة AutoPlay الحالية
        autoplay_status = USBBlocker.get_autoplay_status()

        stats = [
            ("WHITELISTED", str(whitelist_count), Colors.SUCCESS, "Trusted Devices"),
            ("BLACKLISTED", str(blacklist_count), Colors.DANGER, "Known Threats"),
            ("AUTO-BLOCKED", str(auto_count), Colors.WARNING, "Unrecognized Devices"),
            ("LOG ENTRIES", str(audit_count), Colors.SECONDARY, "Security Events"),
        ]
        for i, (title, val, col, sub) in enumerate(stats):
            card = self.create_card(grid_frame, title, val, col, sub)
            card.place(relx=0.02 + (i*0.23), rely=0.05, relwidth=0.21, relheight=0.25)

        # ── بطاقة AutoPlay (قراءة فقط) ───────────────────────────────────
        ap_color = Colors.DANGER  if autoplay_status == "Enabled"  else Colors.SUCCESS
        ap_icon  = "⚠️ ON"        if autoplay_status == "Enabled"  else "🛡️ OFF"
        ap_note  = "Risk: Auto-execution enabled" if autoplay_status == "Enabled" \
                   else "Protected by USB Shield"

        ap_card = tk.Frame(grid_frame, bg=Colors.BG_CARD,
                           highlightbackground=ap_color, highlightthickness=2)
        ap_card.place(relx=0.02, rely=0.38, relwidth=0.44, relheight=0.18)

        ap_inner = tk.Frame(ap_card, bg=Colors.BG_CARD, padx=20, pady=10)
        ap_inner.pack(fill=tk.BOTH, expand=True)

        left = tk.Frame(ap_inner, bg=Colors.BG_CARD)
        left.pack(side=tk.LEFT, fill=tk.Y)

        tk.Label(left, text="🔁  AUTOPLAY STATUS", font=("Segoe UI", 9, "bold"),
                 fg=Colors.TEXT_DIM, bg=Colors.BG_CARD).pack(anchor='w')
        tk.Label(left, text=ap_icon, font=("Segoe UI", 22, "bold"),
                 fg=ap_color, bg=Colors.BG_CARD).pack(anchor='w')
        tk.Label(left, text=ap_note, font=("Segoe UI", 8),
                 fg=Colors.TEXT_DIM, bg=Colors.BG_CARD).pack(anchor='w')

        right = tk.Frame(ap_inner, bg=Colors.BG_CARD)
        right.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))

        badge = tk.Frame(right, bg=Colors.PRIMARY, padx=10, pady=6)
        badge.pack(anchor='center', expand=True)
        tk.Label(badge, text="🔒  Managed by USB Shield",
                 font=("Segoe UI", 8, "bold"),
                 fg=Colors.BG_DARK, bg=Colors.PRIMARY).pack()

        # ── رسم بياني: نشاط آخر 7 أيام ───────────────────────────────────
        try:
            chart_card = tk.Frame(grid_frame, bg=Colors.BG_CARD)
            chart_card.place(relx=0.50, rely=0.38, relwidth=0.48, relheight=0.55)
            tk.Label(chart_card, text="📊  Activity — Last 7 Days",
                     font=("Segoe UI", 9, "bold"), fg=Colors.TEXT_DIM,
                     bg=Colors.BG_CARD).pack(anchor='w', padx=15, pady=(10, 0))

            # جلب بيانات آخر 7 أيام من audit_log
            from datetime import timedelta
            conn2 = get_connection()
            cur2  = conn2.cursor()
            daily = {}
            for i in range(6, -1, -1):
                day = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
                cur2.execute("SELECT COUNT(*) FROM audit_log WHERE timestamp LIKE ?", (f"{day}%",))
                daily[day] = cur2.fetchone()[0]
            conn2.close()

            canvas_h = 100
            canvas = tk.Canvas(chart_card, bg=Colors.BG_CARD,
                                height=canvas_h, highlightthickness=0)
            canvas.pack(fill=tk.X, padx=15, pady=8)
            canvas.update_idletasks()
            cw = canvas.winfo_width() or 300

            days_list  = list(daily.items())
            max_val    = max(v for _, v in days_list) or 1
            bar_w      = max(8, cw // 9)
            gap        = (cw - len(days_list) * bar_w) // (len(days_list) + 1)
            bar_colors = [Colors.DANGER if v > 0 else Colors.SUCCESS for _, v in days_list]

            for idx, ((day, val), bcolor) in enumerate(zip(days_list, bar_colors)):
                x0 = gap + idx * (bar_w + gap)
                x1 = x0 + bar_w
                bh = max(4, int((val / max_val) * (canvas_h - 24)))
                y0 = canvas_h - 18 - bh
                canvas.create_rectangle(x0, y0, x1, canvas_h - 18,
                                        fill=bcolor, outline="")
                if val > 0:
                    canvas.create_text((x0+x1)//2, y0-4, text=str(val),
                                       font=("Segoe UI", 7), fill=Colors.TEXT_DIM)
                label = day[-5:]   # MM-DD
                canvas.create_text((x0+x1)//2, canvas_h-8, text=label,
                                   font=("Segoe UI", 7), fill=Colors.TEXT_DIM)
        except Exception as e:
            print(f"⚠️ Chart error: {e}")

        

    def show_connected_usb(self):
        self.current_view = 'connected_usb'
        self.clear_view()
        header = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40, pady=30)
        header.pack(fill=tk.X)
        tk.Label(header, text="Live Device Monitor", font=("Segoe UI", 24, "bold"), 
                fg=Colors.TEXT_MAIN, bg=Colors.BG_DARK).pack(side=tk.LEFT)
        
        # ✅ الأزرار: SCAN + DEVICE INFO + DELETE
        btn_frame = tk.Frame(header, bg=Colors.BG_DARK)
        btn_frame.pack(side=tk.RIGHT)
        
        HoverButton(btn_frame, text="🔄 SCAN", font=("Segoe UI", 9, "bold"), 
                bg=Colors.PRIMARY, fg=Colors.BG_DARK, hover_bg='#00b8e6', 
                relief='flat', bd=0, padx=20, command=self.show_connected_usb).pack(side=tk.RIGHT, padx=5)
        
        HoverButton(btn_frame, text="ℹ️ DEVICE INFO", font=("Segoe UI", 9, "bold"), 
                bg=Colors.INFO, fg=Colors.BG_DARK, hover_bg='#9900ff', 
                relief='flat', bd=0, padx=20, command=self.show_device_info).pack(side=tk.RIGHT, padx=5)

        HoverButton(btn_frame, text="🗑️ DELETE", font=("Segoe UI", 9, "bold"),
                bg=Colors.DANGER, fg=Colors.TEXT_MAIN, hover_bg='#cc0033',
                relief='flat', bd=0, padx=20,
                command=self.delete_selected_devices).pack(side=tk.RIGHT, padx=5)
        
        tk.Label(self.main_container, 
                text="⚠️ AUTO-BLOCK MODE: All unrecognized devices are blocked", 
                font=("Segoe UI", 9, "bold"), fg=Colors.DANGER, bg=Colors.BG_DARK).pack(pady=(0, 10))

        table_frame = tk.Frame(self.main_container, bg=Colors.BG_CARD, padx=20, pady=20)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=(0, 40))
        
        # ✅ الأعمدة: SELECT + MODEL + SERIAL + CAPACITY + STATUS
        headers = ["SELECT", "MODEL", "SERIAL", "CAPACITY", "STATUS"]
        for c, h in enumerate(headers):
            tk.Label(table_frame, text=h, font=("Segoe UI", 9, "bold"), 
                    fg=Colors.PRIMARY, bg=Colors.BG_CARD, pady=10).grid(row=0, column=c, sticky='nsew')

        # ── 1. الأجهزة المتصلة فعلياً ────────────────────────────────────────
        connected_devices = get_connected_usb_devices()
        connected_fps     = {dev['fingerprint'] for dev in connected_devices}

        # ── 2. كل الأجهزة المصنّفة في القوائم الثلاث ────────────────────────
        wl_devices = get_all_whitelist_devices()
        bl_devices = get_all_blacklist_devices()
        ab_devices = DeviceListManager.get_auto_blocked_devices()

        # ── 3. بناء قاموس شامل fingerprint → device (بدون تكرار) ────────────
        all_known = {}

        for dev in wl_devices:
            fp = dev.get('fingerprint')
            if fp:
                all_known[fp] = {
                    'model':         dev.get('model', 'Unknown'),
                    'serial_number': dev.get('serial_number', 'N/A'),
                    'size_gb':       dev.get('size_gb', 0),
                    'fingerprint':   fp,
                    'vid':           dev.get('vid', 'N/A'),
                    'pid':           dev.get('pid', 'N/A'),
                    'pnp_device_id': None,
                    'is_connected':  False,
                }

        for dev in bl_devices:
            fp = dev.get('fingerprint')
            if fp and fp not in all_known:
                all_known[fp] = {
                    'model':         dev.get('model', 'Unknown'),
                    'serial_number': dev.get('serial_number', 'N/A'),
                    'size_gb':       dev.get('size_gb', 0),
                    'fingerprint':   fp,
                    'vid':           dev.get('vid', 'N/A'),
                    'pid':           dev.get('pid', 'N/A'),
                    'pnp_device_id': None,
                    'is_connected':  False,
                }

        for dev in ab_devices:
            fp = dev.get('fingerprint')
            if fp and fp not in all_known:
                all_known[fp] = {
                    'model':         dev.get('model', 'Unknown'),
                    'serial_number': dev.get('serial_number', 'N/A'),
                    'size_gb':       dev.get('size_gb', 0),
                    'fingerprint':   fp,
                    'vid':           dev.get('vid', 'N/A'),
                    'pid':           dev.get('pid', 'N/A'),
                    'pnp_device_id': dev.get('pnp_device_id', None),
                    'is_connected':  False,
                }

        # ── 4. الأجهزة المتصلة تحصل على بيانات WMI الحديثة وتُزال من all_known ─
        final_list = []
        for dev in connected_devices:
            dev['is_connected'] = True
            all_known.pop(dev['fingerprint'], None)
            final_list.append(dev)

        # أضف الأجهزة غير المتصلة (موجودة في قاعدة البيانات فقط)
        for fp, dev in all_known.items():
            final_list.append(dev)

        # ── عرض الجدول ────────────────────────────────────────────────────────
        self.connected_usb_selection = {}

        if not final_list:
            tk.Label(table_frame, text="📭 No USB devices in the system",
                    font=("Segoe UI", 11), fg=Colors.TEXT_DIM,
                    bg=Colors.BG_CARD).grid(row=1, columnspan=5, pady=50)
        else:
            for r, dev in enumerate(final_list, 1):
                bg_row       = Colors.BG_CARD if r % 2 == 0 else Colors.SIDEBAR
                is_connected = dev.get('is_connected', False)

                if is_connected:
                    model_text = dev.get('model', 'Unknown')[:30]
                    fg_color   = Colors.TEXT_MAIN
                else:
                    model_text = f"🔌 {dev.get('model','Unknown')[:22]} (Disconnected)"
                    fg_color   = Colors.TEXT_DIM

                status = get_device_status(dev['fingerprint'])

                var = tk.BooleanVar()
                self.connected_usb_selection[dev['fingerprint']] = (var, dev)
                tk.Checkbutton(table_frame, variable=var, bg=bg_row,
                               activebackground=bg_row).grid(row=r, column=0, padx=5)

                tk.Label(table_frame, text=model_text, fg=fg_color,
                         bg=bg_row).grid(row=r, column=1, sticky='nsew', ipady=5, padx=5)
                tk.Label(table_frame, text=dev.get('serial_number','N/A')[:20], fg=fg_color,
                         bg=bg_row).grid(row=r, column=2, sticky='nsew', padx=5)
                tk.Label(table_frame, text=f"{dev.get('size_gb', 0)} GB", fg=fg_color,
                         bg=bg_row).grid(row=r, column=3, sticky='nsew', padx=5)

                if status == "Whitelisted":
                    s_color, s_text = Colors.SUCCESS, "✅ AUTHORIZED"
                elif status == "Blacklisted":
                    s_color, s_text = Colors.DANGER,  "⛔ BLACKLISTED"
                else:
                    s_color, s_text = Colors.WARNING,  "🔒 BLOCKED"
                tk.Label(table_frame, text=s_text, fg=s_color, bg=bg_row,
                         font=("Segoe UI", 9, "bold")).grid(row=r, column=4, sticky='nsew', padx=5)
                        
        for i in range(5): 
            table_frame.grid_columnconfigure(i, weight=1)


    def show_device_info(self):
        """عرض معلومات الجهاز المحدد"""
        # ✅ التحقق من وجود تحديد
        if not hasattr(self, 'connected_usb_selection') or not self.connected_usb_selection:
            messagebox.showwarning("Warning", "⚠️ No devices available!")
            return
        
        selected = [fp for fp, (var, dev) in self.connected_usb_selection.items() if var.get()]
        
        if not selected:
            messagebox.showwarning("Warning", "⚠️ Please select a device first!")
            return
        
        if len(selected) > 1:
            messagebox.showwarning("Warning", "⚠️ Please select only ONE device!")
            return
        
        # ✅ التحقق من كلمة المرور
        if not self.verify_user_password("View Device Information"):
            return
        
        # ✅ جلب معلومات الجهاز
        fingerprint = selected[0]
        _, device = self.connected_usb_selection[fingerprint]
        
        # ✅ إنشاء نافذة منبثقة في المنتصف
        info_dialog = tk.Toplevel(self.root)
        info_dialog.title("📱 Device Information")
        info_dialog.configure(bg=Colors.BG_DARK)
        info_dialog.transient(self.root)
        info_dialog.grab_set()
        
        # ✅ توسيط النافذة
        self._center_window(info_dialog, 650, 550)
        
        # ✅ المحتوى
        header_frame = tk.Frame(info_dialog, bg=Colors.BG_CARD, pady=20)
        header_frame.pack(fill=tk.X, padx=20, pady=(20, 10))
        
        tk.Label(header_frame, text="🔌 USB DEVICE DETAILS", 
                font=("Segoe UI", 16, "bold"), 
                fg=Colors.PRIMARY, bg=Colors.BG_CARD).pack()
        tk.Label(header_frame, text=device['model'][:50], 
                font=("Segoe UI", 11), 
                fg=Colors.TEXT_MAIN, bg=Colors.BG_CARD).pack(pady=5)
        
        # ✅ معلومات الجهاز
        content_frame = tk.Frame(info_dialog, bg=Colors.BG_CARD, padx=30, pady=20)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # ✅ قائمة المعلومات
        info_items = [
            ("🔹 Fingerprint:", device.get('fingerprint', 'N/A')),
            ("🔹 Model:", device.get('model', 'N/A')),
            ("🔹 Serial Number:", device.get('serial_number', 'N/A')),
            ("🔹 VID (Vendor ID):", device.get('vid', 'N/A')),
            ("🔹 PID (Product ID):", device.get('pid', 'N/A')),
            ("🔹 Capacity:", f"{device.get('size_gb', 0)} GB"),
            ("🔹 Device ID:", device.get('device_id', 'N/A')),
            ("🔹 PnP Device ID:", device.get('pnp_device_id', 'N/A')),
            ("🔹 Status:", get_device_status(device.get('fingerprint', ''))),
        ]
        
        for label, value in info_items:
            frame = tk.Frame(content_frame, bg=Colors.BG_CARD)
            frame.pack(fill=tk.X, pady=5)
            
            tk.Label(frame, text=label, font=("Segoe UI", 9, "bold"), 
                    fg=Colors.PRIMARY, bg=Colors.BG_CARD, 
                    width=25, anchor='w').pack(side=tk.LEFT)
            
            # ✅ قيمة قابلة للتحديد والنسخ
            val_label = tk.Label(frame, text=str(value), font=("Segoe UI", 9), 
                                fg=Colors.TEXT_MAIN, bg=Colors.BG_DARK, 
                                padx=10, pady=5, anchor='w')
            val_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # ✅ زر الإغلاق
        btn_frame = tk.Frame(info_dialog, bg=Colors.BG_DARK, pady=20)
        btn_frame.pack(fill=tk.X, padx=20)
        
        HoverButton(btn_frame, text="❌ CLOSE", font=("Segoe UI", 10, "bold"), 
                bg=Colors.DANGER, fg=Colors.TEXT_MAIN, hover_bg='#cc0044', 
                relief='flat', bd=0, padx=40, command=info_dialog.destroy).pack()
        
        # ✅ منع التفاعل مع النافذة الرئيسية
        info_dialog.protocol("WM_DELETE_WINDOW", info_dialog.destroy)

    def delete_selected_devices(self):
        """
        حذف كامل ونهائي للجهاز من:
          - قاعدة البيانات (whitelist, blacklist, auto_blocked)
          - Registry policy (deny list)
          - Device Manager (تفعيل الجهاز إذا كان معطّلاً)
          - ذاكرة USBMonitor (يُعامَل كجهاز جديد عند إعادة التوصيل)
        """
        if not hasattr(self, 'connected_usb_selection') or not self.connected_usb_selection:
            messagebox.showwarning("Warning", "⚠️ No devices available!")
            return

        selected = [(fp, dev) for fp, (var, dev) in self.connected_usb_selection.items() if var.get()]
        if not selected:
            messagebox.showwarning("Warning", "⚠️ Please select at least one device first!")
            return

        if not self.verify_user_password(f"Delete {len(selected)} device(s) permanently"):
            return

        names = "\n".join(f"  • {dev.get('model', 'Unknown')[:45]}" for _, dev in selected)
        confirm = messagebox.askyesno(
            "⚠️  Confirm Permanent Delete",
            f"This will PERMANENTLY delete {len(selected)} device(s) from ALL lists:\n\n"
            f"{names}\n\n"
            "• Removed from: Whitelist, Blacklist, Auto-Blocked\n"
            "• Block removed from Windows Registry\n"
            "• Device re-enabled if it was disabled\n"
            "• On next connection → treated as a brand-new device\n\n"
            "This cannot be undone. Continue?",
            icon='warning'
        )
        if not confirm:
            return

        deleted = 0
        try:
            conn = get_connection()
            cursor = conn.cursor()

            for fp, dev in selected:
                model   = dev.get('model', 'Unknown')
                pnp_id  = dev.get('pnp_device_id') or dev.get('pnp_id', '')
                serial  = dev.get('serial_number', 'N/A')
                vid     = dev.get('vid', 'N/A')
                pid     = dev.get('pid', 'N/A')
                size_gb = dev.get('size_gb', 0)

                # ── 1. تحديد القوائم التي كان فيها الجهاز ───────────────────
                lists_found = []
                cursor.execute("SELECT id FROM whitelist    WHERE fingerprint = ?", (fp,))
                if cursor.fetchone(): lists_found.append("Whitelist")
                cursor.execute("SELECT id FROM blacklist    WHERE fingerprint = ?", (fp,))
                if cursor.fetchone(): lists_found.append("Blacklist")
                cursor.execute("SELECT id FROM auto_blocked WHERE fingerprint = ?", (fp,))
                if cursor.fetchone(): lists_found.append("Auto-Blocked")

                # ── إذا لم نجد pnp_id في الذاكرة، نبحث في قاعدة البيانات ──
                if not pnp_id:
                    cursor.execute("SELECT pnp_device_id FROM auto_blocked WHERE fingerprint = ?", (fp,))
                    row = cursor.fetchone()
                    if row and row[0]:
                        pnp_id = row[0]
                if not pnp_id:
                    try:
                        cursor.execute("SELECT pnp_device_id FROM blacklist WHERE fingerprint = ?", (fp,))
                        row = cursor.fetchone()
                        if row and row[0]:
                            pnp_id = row[0]
                    except Exception:
                        pass

                lists_str = ", ".join(lists_found) if lists_found else "None"

                # ── 2. حذف من كل الجداول ──────────────────────────────────
                cursor.execute("DELETE FROM whitelist    WHERE fingerprint = ?", (fp,))
                cursor.execute("DELETE FROM blacklist    WHERE fingerprint = ?", (fp,))
                cursor.execute("DELETE FROM auto_blocked WHERE fingerprint = ?", (fp,))

                # ── 3. إزالة الحظر من Registry + تفعيل الجهاز ────────────
                # يجب أن يحدث بعد الحذف من قاعدة البيانات حتى لا يُعاد حظره
                system_cleaned = False
                if pnp_id:
                    try:
                        # حذف من Registry deny list
                        from usb_blocker import remove_from_deny_list
                        remove_from_deny_list(pnp_id)

                        # تعطيل السياسة الاسترجاعية مؤقتاً
                        try:
                            winreg.CreateKey(winreg.HKEY_LOCAL_MACHINE,
                                r"SOFTWARE\Policies\Microsoft\Windows\DeviceInstall\Restrictions")
                            with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                                r"SOFTWARE\Policies\Microsoft\Windows\DeviceInstall\Restrictions",
                                0, winreg.KEY_SET_VALUE) as k:
                                winreg.SetValueEx(k, "DenyDeviceIDsRetroactive", 0, winreg.REG_DWORD, 0)
                        except Exception:
                            pass

                        # تفعيل الجهاز عبر PowerShell (يشمل كل عقد الجهاز بالـ serial)
                        serial_clean = pnp_id.split('\\')[-1].split('&')[0].strip() if '\\' in pnp_id else serial
                        if serial_clean and serial_clean != 'N/A':
                            ps_script = f'''
$serial = "{serial_clean}"
$devices = Get-PnpDevice | Where-Object {{ $_.InstanceId -like "*$serial*" }}
foreach ($dev in $devices) {{
    Enable-PnpDevice -InstanceId $dev.InstanceId -Confirm:$false -ErrorAction SilentlyContinue
}}
Write-Host "Enabled $($devices.Count) device(s) for serial $serial"
'''
                            import subprocess as _sp
                            _sp.run(['powershell', '-NonInteractive', '-ExecutionPolicy', 'Bypass',
                                     '-Command', ps_script],
                                    capture_output=True, text=True, timeout=20)

                        # إعادة السياسة الاسترجاعية
                        try:
                            with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                                r"SOFTWARE\Policies\Microsoft\Windows\DeviceInstall\Restrictions",
                                0, winreg.KEY_SET_VALUE) as k:
                                winreg.SetValueEx(k, "DenyDeviceIDsRetroactive", 0, winreg.REG_DWORD, 1)
                        except Exception:
                            pass

                        # pnputil scan
                        _sp.run('pnputil /scan-devices', shell=True,
                                capture_output=True, text=True, timeout=10)

                        system_cleaned = True
                        print(f"🔓 System-level block removed for: {model}")
                    except Exception as e:
                        print(f"⚠️ System cleanup warning: {e}")

                # ── 4. إزالة من ذاكرة USBMonitor ─────────────────────────
                if pnp_id and hasattr(self.usb_monitor, 'known_pnp_ids'):
                    self.usb_monitor.known_pnp_ids.discard(pnp_id)

                # ── 5. تسجيل في Audit Logs ──────────────────────────────
                log_event(
                    event_type         = "DEVICE_PERMANENTLY_DELETED",
                    device_fingerprint = fp,
                    device_model       = model,
                    result             = "Success",
                    user               = self.username,
                    details            = (
                        f"Permanently deleted from: [{lists_str}]. "
                        f"Serial: {serial} | VID: {vid} | PID: {pid} | Size: {size_gb} GB | "
                        f"PnP: {pnp_id or 'N/A'} | "
                        f"System block removed: {'Yes' if system_cleaned else 'No (pnp_id unavailable)'}. "
                        f"Device will be treated as brand-new on next connection."
                    )
                )
                deleted += 1
                print(f"🗑️ Permanently deleted: {model} | FP: {fp[:14]}... | Was in: [{lists_str}] | System cleaned: {system_cleaned}")

            conn.commit()
            conn.close()

        except Exception as e:
            messagebox.showerror("Delete Error", f"❌ Failed to delete:\n{e}")
            return

        sys_note = "\n• Block removed from Windows and device re-enabled" if deleted > 0 else ""
        messagebox.showinfo(
            "✅  Delete Complete",
            f"{deleted} device(s) permanently removed from all lists.{sys_note}\n\n"
            "On next connection they will be treated as new devices\n"
            "and added to Auto-Blocked automatically."
        )
        self.show_connected_usb()

    def show_whitelist(self):
        self.current_view = 'whitelist'
        self.clear_view()
        header = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40, pady=30)
        header.pack(fill=tk.X)
        tk.Label(header, text="Trusted Devices", font=("Segoe UI", 24, "bold"), fg=Colors.SUCCESS, bg=Colors.BG_DARK).pack(side=tk.LEFT)
        
        # ✅ الأزرار الثلاثة
        btn_frame = tk.Frame(header, bg=Colors.BG_DARK)
        btn_frame.pack(side=tk.RIGHT)
        
        HoverButton(btn_frame, text="➕ Add Device", bg=Colors.SUCCESS, fg=Colors.BG_DARK, hover_bg='#00cc7a', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, 
                   command=self.add_whitelist_device).pack(side=tk.LEFT, padx=5)
        
        HoverButton(btn_frame, text="🗑️ Remove", bg=Colors.DANGER, fg=Colors.TEXT_MAIN, hover_bg='#cc0044', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, 
                   command=self.remove_whitelist_device).pack(side=tk.LEFT, padx=5)
        
        HoverButton(btn_frame, text="🔄 Refresh", bg=Colors.PRIMARY, fg=Colors.BG_DARK, hover_bg='#00b8e6', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, 
                   command=self.show_whitelist).pack(side=tk.LEFT, padx=5)

        # ── شريط البحث ────────────────────────────────────────────────────
        search_frame = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40)
        search_frame.pack(fill=tk.X, pady=(0, 8))
        tk.Label(search_frame, text="🔍", font=("Segoe UI", 11),
                 fg=Colors.TEXT_DIM, bg=Colors.BG_DARK).pack(side=tk.LEFT, padx=(0, 6))
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var,
                                font=("Segoe UI", 10), bg=Colors.BG_CARD,
                                fg=Colors.TEXT_MAIN, insertbackground='white',
                                relief='flat', bd=6)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Label(search_frame, text="Search by model or serial",
                 font=("Segoe UI", 8), fg=Colors.TEXT_DIM, bg=Colors.BG_DARK).pack(side=tk.LEFT, padx=8)

        table_frame = tk.Frame(self.main_container, bg=Colors.BG_CARD, padx=20, pady=20)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=(0, 40))
        headers = ["ID", "MODEL", "SERIAL", "VID", "PID", "SIZE", "ADDED BY", "DATE"]
        for c, h in enumerate(headers):
            tk.Label(table_frame, text=h, font=("Segoe UI", 9, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_CARD, pady=10).grid(row=0, column=c, sticky='nsew')

        all_devices = get_all_whitelist_devices()
        row_widgets = []

        def _render_rows(devices):
            for w in row_widgets:
                for lbl in w: lbl.destroy()
            row_widgets.clear()
            if not devices:
                lbl = tk.Label(table_frame, text="📭 No whitelisted devices", font=("Segoe UI", 11), fg=Colors.TEXT_DIM, bg=Colors.BG_CARD)
                lbl.grid(row=1, columnspan=8, pady=50)
                row_widgets.append([lbl])
            else:
                for r, dev in enumerate(devices, 1):
                    bg_row = Colors.BG_CARD if r % 2 == 0 else Colors.SIDEBAR
                    values = [dev.get('id',''), dev.get('model','')[:25], dev.get('serial_number','')[:18],
                              dev.get('vid',''), dev.get('pid',''), str(dev.get('size_gb','')),
                              dev.get('added_by','')[:12], dev.get('date_added','')[:19] if dev.get('date_added') else '']
                    lbls = []
                    for c, val in enumerate(values):
                        lbl = tk.Label(table_frame, text=str(val), fg=Colors.TEXT_MAIN, bg=bg_row)
                        lbl.grid(row=r, column=c, sticky='nsew', ipady=5, padx=5)
                        lbls.append(lbl)
                    row_widgets.append(lbls)

        def _on_search(*_):
            q = search_var.get().strip().lower()
            filtered = [d for d in all_devices
                        if q in (d.get('model','') or '').lower()
                        or q in (d.get('serial_number','') or '').lower()] if q else all_devices
            _render_rows(filtered)

        search_var.trace_add('write', _on_search)
        _render_rows(all_devices)
        for i in range(8): table_frame.grid_columnconfigure(i, weight=1)

    def show_blacklist(self):
        self.current_view = 'blacklist'
        self.clear_view()
        header = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40, pady=30)
        header.pack(fill=tk.X)
        tk.Label(header, text="Blocked Devices", font=("Segoe UI", 24, "bold"), fg=Colors.DANGER, bg=Colors.BG_DARK).pack(side=tk.LEFT)
        btn_frame = tk.Frame(header, bg=Colors.BG_DARK)
        btn_frame.pack(side=tk.RIGHT)
        HoverButton(btn_frame, text="⛔ Add Device", bg=Colors.DANGER, fg=Colors.TEXT_MAIN, hover_bg='#cc0044', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, command=self.add_blacklist_device).pack(side=tk.LEFT, padx=5)
        HoverButton(btn_frame, text="✅ Unblock", bg=Colors.SUCCESS, fg=Colors.BG_DARK, hover_bg='#00cc7a', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, command=self.remove_blacklist_device).pack(side=tk.LEFT, padx=5)
        HoverButton(btn_frame, text="🔄 Refresh", bg=Colors.PRIMARY, fg=Colors.BG_DARK, hover_bg='#00b8e6', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, command=self.show_blacklist).pack(side=tk.LEFT, padx=5)

        # ── شريط البحث ────────────────────────────────────────────────────
        bl_search_frame = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40)
        bl_search_frame.pack(fill=tk.X, pady=(0, 8))
        tk.Label(bl_search_frame, text="🔍", font=("Segoe UI", 11),
                 fg=Colors.TEXT_DIM, bg=Colors.BG_DARK).pack(side=tk.LEFT, padx=(0, 6))
        bl_search_var = tk.StringVar()
        tk.Entry(bl_search_frame, textvariable=bl_search_var,
                 font=("Segoe UI", 10), bg=Colors.BG_CARD,
                 fg=Colors.TEXT_MAIN, insertbackground='white',
                 relief='flat', bd=6).pack(side=tk.LEFT, fill=tk.X, expand=True)

        table_frame = tk.Frame(self.main_container, bg=Colors.BG_CARD, padx=20, pady=20)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=(0, 40))
        headers = ["ID", "MODEL", "SERIAL", "REASON", "BLOCKED BY", "DATE"]
        for c, h in enumerate(headers):
            tk.Label(table_frame, text=h, font=("Segoe UI", 9, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_CARD, pady=10).grid(row=0, column=c, sticky='nsew')

        all_bl = get_all_blacklist_devices()
        bl_rows = []

        def _render_bl(devices):
            for w in bl_rows:
                for lbl in w: lbl.destroy()
            bl_rows.clear()
            if not devices:
                lbl = tk.Label(table_frame, text="📭 No blocked devices", font=("Segoe UI", 11), fg=Colors.TEXT_DIM, bg=Colors.BG_CARD)
                lbl.grid(row=1, columnspan=6, pady=50); bl_rows.append([lbl])
            else:
                for r, dev in enumerate(devices, 1):
                    bg_row = Colors.BG_CARD if r % 2 == 0 else Colors.SIDEBAR
                    values = [dev.get('id',''), dev.get('model','')[:25], dev.get('serial_number','')[:18],
                              dev.get('block_reason','N/A')[:20], dev.get('blocked_by','')[:12],
                              dev.get('date_blocked','')[:19] if dev.get('date_blocked') else '']
                    lbls = []
                    for c, val in enumerate(values):
                        lbl = tk.Label(table_frame, text=str(val), fg=Colors.TEXT_MAIN, bg=bg_row)
                        lbl.grid(row=r, column=c, sticky='nsew', ipady=5, padx=5); lbls.append(lbl)
                    bl_rows.append(lbls)

        def _on_bl_search(*_):
            q = bl_search_var.get().strip().lower()
            filtered = [d for d in all_bl if q in (d.get('model','') or '').lower()
                        or q in (d.get('serial_number','') or '').lower()] if q else all_bl
            _render_bl(filtered)

        bl_search_var.trace_add('write', _on_bl_search)
        _render_bl(all_bl)
        for i in range(6): table_frame.grid_columnconfigure(i, weight=1)

    def show_auto_blocked(self):
        """عرض قائمة الأجهزة المحظورة تلقائياً"""
        self.current_view = 'auto_blocked'
        self.clear_view()
        header = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40, pady=30)
        header.pack(fill=tk.X)
        tk.Label(header, text="Auto-Blocked Devices", font=("Segoe UI", 24, "bold"), fg=Colors.WARNING, bg=Colors.BG_DARK).pack(side=tk.LEFT)
        btn_frame = tk.Frame(header, bg=Colors.BG_DARK)
        btn_frame.pack(side=tk.RIGHT)
        HoverButton(btn_frame, text="✓ Trust", bg=Colors.SUCCESS, fg=Colors.BG_DARK, hover_bg='#00cc7a', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, command=lambda: self.move_from_auto('whitelist')).pack(side=tk.LEFT, padx=5)
        HoverButton(btn_frame, text="⛔ Block Permanently", bg=Colors.DANGER, fg=Colors.TEXT_MAIN, hover_bg='#cc0044', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, command=lambda: self.move_from_auto('blacklist')).pack(side=tk.LEFT, padx=5)
        HoverButton(btn_frame, text="🔄 Refresh", bg=Colors.PRIMARY, fg=Colors.BG_DARK, hover_bg='#00b8e6', 
                   font=("Segoe UI", 9, "bold"), relief='flat', bd=0, padx=15, command=self.show_auto_blocked).pack(side=tk.LEFT, padx=5)

        table_frame = tk.Frame(self.main_container, bg=Colors.BG_CARD, padx=20, pady=20)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=(0, 40))
        headers = ["SELECT", "MODEL", "SERIAL", "FIRST SEEN", "LAST SEEN", "REASON"]
        for c, h in enumerate(headers):
            tk.Label(table_frame, text=h, font=("Segoe UI", 9, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_CARD, pady=10).grid(row=0, column=c, sticky='nsew')

        devices = DeviceListManager.get_auto_blocked_devices()
        self.auto_blocked_selection = {}
        
        if not devices:
            tk.Label(table_frame, text="📭 No auto-blocked devices", font=("Segoe UI", 11), fg=Colors.TEXT_DIM, bg=Colors.BG_CARD).grid(row=1, columnspan=6, pady=50)
        else:
            for r, dev in enumerate(devices, 1):
                bg_row = Colors.BG_CARD if r % 2 == 0 else Colors.SIDEBAR
                var = tk.BooleanVar()
                self.auto_blocked_selection[dev['fingerprint']] = var
                tk.Checkbutton(table_frame, variable=var, bg=bg_row, activebackground=bg_row).grid(row=r, column=0, padx=5)
                values = [dev['model'][:25], dev['serial_number'][:18], 
                         dev['first_seen'][:19] if dev['first_seen'] else '', 
                         dev['last_seen'][:19] if dev['last_seen'] else '',
                         dev['block_reason'][:25]]
                for c, val in enumerate(values, start=1):
                    tk.Label(table_frame, text=str(val), fg=Colors.TEXT_MAIN, bg=bg_row).grid(row=r, column=c, sticky='nsew', ipady=5, padx=5)
        for i in range(6): table_frame.grid_columnconfigure(i, weight=1)

    def show_audit_logs(self):
        """عرض سجل الأحداث كجدول Treeview ملون — كل نوع حدث له لون مختلف."""
        self.current_view = 'audit_logs'
        self.clear_view()

        # ── الهيدر ────────────────────────────────────────────────────────
        header = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40, pady=25)
        header.pack(fill=tk.X)

        tk.Label(header, text="📜  Security Audit Logs",
                 font=("Segoe UI", 22, "bold"),
                 fg=Colors.WARNING, bg=Colors.BG_DARK).pack(side=tk.LEFT)

        btn_frame = tk.Frame(header, bg=Colors.BG_DARK)
        btn_frame.pack(side=tk.RIGHT)

        HoverButton(btn_frame, text="📥  Export",
                    bg=Colors.SUCCESS, fg=Colors.BG_DARK,
                    hover_bg='#00cc7a', font=("Segoe UI", 9, "bold"),
                    relief='flat', bd=0, padx=18,
                    command=self.export_audit_logs).pack(side=tk.LEFT, padx=4)

        HoverButton(btn_frame, text="🗑️  Clear",
                    bg=Colors.DANGER, fg=Colors.TEXT_MAIN,
                    hover_bg='#cc0044', font=("Segoe UI", 9, "bold"),
                    relief='flat', bd=0, padx=18,
                    command=self.clear_audit_logs).pack(side=tk.LEFT, padx=4)

        HoverButton(btn_frame, text="🔄  Refresh",
                    bg=Colors.BG_CARD, fg=Colors.PRIMARY,
                    hover_bg=Colors.BORDER, font=("Segoe UI", 9, "bold"),
                    relief='flat', bd=0, padx=18,
                    command=self.show_audit_logs).pack(side=tk.LEFT, padx=4)

        # ── حقل بحث ──────────────────────────────────────────────────────
        search_bar = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40)
        search_bar.pack(fill=tk.X, pady=(0, 8))

        tk.Label(search_bar, text="🔍", font=("Segoe UI", 11),
                 fg=Colors.TEXT_DIM, bg=Colors.BG_DARK).pack(side=tk.LEFT, padx=(0, 6))

        self._log_search_var = tk.StringVar()
        search_entry = tk.Entry(search_bar, textvariable=self._log_search_var,
                                font=("Segoe UI", 9), bg=Colors.BG_CARD,
                                fg=Colors.TEXT_MAIN, relief='flat',
                                insertbackground=Colors.PRIMARY)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True,
                          ipady=6, padx=(0, 10))
        tk.Label(search_bar, text="Filter by event type or user",
                 font=("Segoe UI", 8), fg=Colors.TEXT_DIM,
                 bg=Colors.BG_DARK).pack(side=tk.LEFT)

        # ── إطار الجدول ──────────────────────────────────────────────────
        table_wrap = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40)
        table_wrap.pack(fill=tk.BOTH, expand=True, pady=(0, 30))

        # Treeview style
        style = ttk.Style()
        style.theme_use('default')
        style.configure("Log.Treeview",
                         background=Colors.BG_CARD,
                         foreground=Colors.TEXT_MAIN,
                         fieldbackground=Colors.BG_CARD,
                         rowheight=30,
                         font=("Segoe UI", 9),
                         borderwidth=0)
        style.configure("Log.Treeview.Heading",
                         background=Colors.SIDEBAR,
                         foreground=Colors.PRIMARY,
                         font=("Segoe UI", 9, "bold"),
                         relief='flat')
        style.map("Log.Treeview",
                  background=[('selected', Colors.PRIMARY)],
                  foreground=[('selected', Colors.BG_DARK)])

        cols = ("timestamp", "event_type", "user", "result", "device", "details")
        col_names = ("Timestamp", "Event Type", "User", "Result", "Device", "Details")
        col_widths = (145, 220, 80, 80, 130, 280)

        tree_frame = tk.Frame(table_wrap, bg=Colors.BG_CARD)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        tree = ttk.Treeview(tree_frame, columns=cols, show='headings',
                             style="Log.Treeview", selectmode='browse')

        for col, name, w in zip(cols, col_names, col_widths):
            tree.heading(col, text=name)
            tree.column(col, width=w, minwidth=60, anchor='w')

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side=tk.RIGHT,  fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(side=tk.LEFT,  fill=tk.BOTH, expand=True)

        # ── ألوان الصفوف حسب نوع الحدث ───────────────────────────────────
        tree.tag_configure('blocked',   background='#1a0812', foreground='#ff6b8a')
        tree.tag_configure('allowed',   background='#071a10', foreground='#00e68a')
        tree.tag_configure('autoplay',  background='#0d1020', foreground='#00d1ff')
        tree.tag_configure('auth',      background='#1a1208', foreground='#ffcc44')
        tree.tag_configure('system',    background='#0f0f18', foreground='#aa88ff')
        tree.tag_configure('deleted',   background='#1a0f00', foreground='#ff8800')
        tree.tag_configure('default',   background=Colors.BG_CARD, foreground=Colors.TEXT_MAIN)

        # ── تحميل البيانات ────────────────────────────────────────────────
        all_logs = []
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 500')
            all_logs = [dict(row) for row in cursor.fetchall()]
            conn.close()
        except Exception as e:
            print(f"Error loading logs: {e}")

        def _get_tag(event_type: str) -> str:
            et = (event_type or "").upper()
            if "PERMANENTLY_DELETED" in et:
                return 'deleted'
            if any(k in et for k in ("BLOCK", "DENY", "FAILED")):
                return 'blocked'
            if any(k in et for k in ("ALLOW", "WHITELIST")):
                return 'allowed'
            if "AUTOPLAY" in et:
                return 'autoplay'
            if any(k in et for k in ("AUTH", "LOGIN", "PASSWORD")):
                return 'auth'
            if any(k in et for k in ("DATABASE", "SYSTEM", "STARTUP")):
                return 'system'
            return 'default'

        def _populate(filter_text=""):
            tree.delete(*tree.get_children())
            ft = filter_text.lower().strip()
            for log in all_logs:
                et   = log.get('event_type', '')
                user = log.get('user', '')
                if ft and ft not in et.lower() and ft not in user.lower():
                    continue
                tag = _get_tag(et)
                tree.insert('', 'end', values=(
                    log.get('timestamp', ''),
                    et,
                    user,
                    log.get('result', ''),
                    log.get('device_model', '') or '—',
                    log.get('details', '') or '—',
                ), tags=(tag,))

        _populate()

        # ── ربط حقل البحث بالجدول ────────────────────────────────────────
        def _on_search(*_):
            _populate(self._log_search_var.get())

        self._log_search_var.trace_add('write', _on_search)

        # ── شريط الحالة ──────────────────────────────────────────────────
        tk.Label(table_wrap,
                 text=f"Showing last {len(all_logs)} entries   "
                      "■ Blocked  ■ Allowed  ■ AutoPlay  ■ Auth  ■ System  ■ Deleted",
                 font=("Segoe UI", 8), fg=Colors.TEXT_DIM,
                 bg=Colors.BG_DARK).pack(anchor='w', pady=(6, 0))

    def show_settings(self):
        try:
            self.clear_view()
            
            header = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40, pady=30)
            header.pack(fill=tk.X)
            tk.Label(header, text="⚙️ Security Configuration", font=("Segoe UI", 24, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_DARK).pack(side=tk.LEFT)
            
            content_frame = tk.Frame(self.main_container, bg=Colors.BG_DARK, padx=40)
            content_frame.pack(fill=tk.BOTH, expand=True)
            
            banner_bg = '#1a0d12'
            banner = tk.Frame(content_frame, bg=banner_bg, highlightbackground=Colors.DANGER, highlightthickness=1)
            banner.pack(fill=tk.X, pady=(0, 30))
            
            tk.Label(banner, text="🛡️ PROTECTION MODE: DEFAULT DENY (MAXIMUM SECURITY)", font=("Segoe UI", 11, "bold"), fg=Colors.DANGER, bg=banner_bg).pack(pady=(15, 5))
            tk.Label(banner, text="Any unrecognized USB device is automatically blocked. Only whitelisted devices are allowed.", font=("Segoe UI", 9), fg=Colors.TEXT_MAIN, bg=banner_bg).pack(pady=(0, 15))

            cards_container = tk.Frame(content_frame, bg=Colors.BG_DARK)
            cards_container.pack(fill=tk.BOTH, expand=True)

            user_card = tk.Frame(cards_container, bg=Colors.BG_CARD, highlightbackground=Colors.BORDER, highlightthickness=1)
            user_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
             
            tk.Label(user_card, text="👤 Account Profile", font=("Segoe UI", 14, "bold"), fg=Colors.SUCCESS, bg=Colors.BG_CARD).pack(anchor='w', padx=30, pady=(30, 20))
            
            u_grid = tk.Frame(user_card, bg=Colors.BG_CARD, padx=30)
            u_grid.pack(fill=tk.X)
            
            tk.Label(u_grid, text="Current Username:", font=("Segoe UI", 10), bg=Colors.BG_CARD, fg=Colors.TEXT_DIM).grid(row=0, column=0, sticky='w', pady=(0, 15))
            tk.Label(u_grid, text=self.username.upper(), font=("Segoe UI", 11, "bold"), bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN).grid(row=0, column=1, sticky='w', pady=(0, 15), padx=20)
            
            tk.Label(u_grid, text="New Username:", font=("Segoe UI", 10), bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN).grid(row=1, column=0, sticky='w', pady=10)
            self.new_username = tk.Entry(u_grid, font=("Segoe UI", 11), bg=Colors.BG_DARK, fg=Colors.TEXT_MAIN, insertbackground='white', relief='flat', width=25)
            self.new_username.grid(row=1, column=1, pady=10, padx=20, ipady=6)
             
            tk.Label(u_grid, text="Admin Password:", font=("Segoe UI", 10), bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN).grid(row=2, column=0, sticky='w', pady=10)
            self.user_pwd = tk.Entry(u_grid, show="●", font=("Segoe UI", 11), bg=Colors.BG_DARK, fg=Colors.TEXT_MAIN, insertbackground='white', relief='flat', width=25)
            self.user_pwd.grid(row=2, column=1, pady=10, padx=20, ipady=6)
             
            self.new_username.bind('<Return>', lambda e: self.user_pwd.focus())
            self.user_pwd.bind('<Return>', lambda e: self.change_username_action())
            
            HoverButton(user_card, text="Apply Changes", bg=Colors.SUCCESS, fg=Colors.BG_DARK, hover_bg='#00cc7a', 
                    font=("Segoe UI", 10, "bold"), relief='flat', bd=0, padx=25, pady=8, 
                    command=self.change_username_action).pack(anchor='e', padx=50, pady=(30, 20))

            pwd_card = tk.Frame(cards_container, bg=Colors.BG_CARD, highlightbackground=Colors.BORDER, highlightthickness=1)
            pwd_card.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
            
            tk.Label(pwd_card, text="🔐 Security Settings", font=("Segoe UI", 14, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_CARD).pack(anchor='w', padx=30, pady=(30, 20))
            
            p_grid = tk.Frame(pwd_card, bg=Colors.BG_CARD, padx=30)
            p_grid.pack(fill=tk.X)
            
            tk.Label(p_grid, text="Current Password:", font=("Segoe UI", 10), bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN).grid(row=0, column=0, sticky='w', pady=10)
            self.current_pwd = tk.Entry(p_grid, show="●", font=("Segoe UI", 11), bg=Colors.BG_DARK, fg=Colors.TEXT_MAIN, insertbackground='white', relief='flat', width=25)
            self.current_pwd.grid(row=0, column=1, pady=10, padx=20, ipady=6)
             
            tk.Label(p_grid, text="New Password:", font=("Segoe UI", 10), bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN).grid(row=1, column=0, sticky='w', pady=10)
            self.new_pwd = tk.Entry(p_grid, show="●", font=("Segoe UI", 11), bg=Colors.BG_DARK, fg=Colors.TEXT_MAIN, insertbackground='white', relief='flat', width=25)
            self.new_pwd.grid(row=1, column=1, pady=10, padx=20, ipady=6)
             
            tk.Label(p_grid, text="Confirm New:", font=("Segoe UI", 10), bg=Colors.BG_CARD, fg=Colors.TEXT_MAIN).grid(row=2, column=0, sticky='w', pady=10)
            self.confirm_pwd = tk.Entry(p_grid, show="●", font=("Segoe UI", 11), bg=Colors.BG_DARK, fg=Colors.TEXT_MAIN, insertbackground='white', relief='flat', width=25)
            self.confirm_pwd.grid(row=2, column=1, pady=10, padx=20, ipady=6)
             
            self.current_pwd.bind('<Return>', lambda e: self.new_pwd.focus())
            self.new_pwd.bind('<Return>', lambda e: self.confirm_pwd.focus())
            self.confirm_pwd.bind('<Return>', lambda e: self.change_password_action())
            
            HoverButton(pwd_card, text="Update Password", bg=Colors.PRIMARY, fg=Colors.BG_DARK, hover_bg='#00b8e6', 
                    font=("Segoe UI", 10, "bold"), relief='flat', bd=0, padx=25, pady=8, 
                    command=self.change_password_action).pack(anchor='e', padx=50, pady=(30, 20))

        except Exception as e:
            print(f"Error in show_settings: {e}")
            messagebox.showerror("UI Error", f"Could not load settings: {e}")

    def change_username_action(self):
        new_user = self.new_username.get().strip()
        pwd = self.user_pwd.get()
        
        if not new_user or not pwd:
            messagebox.showwarning("Warning", "Please fill all fields!")
            return
            
        from auth_manager import is_valid_username
        if not is_valid_username(new_user):
            messagebox.showerror("Invalid Username", "❌ اسم المستخدم غير صالح!\n\nيجب أن يكون:\n- من 3 إلى 20 حرف\n- أحرف إنجليزية وأرقام فقط\n- بدون مسافات أو رموز خاصة")
            return
        
        if change_username(self.username, new_user, pwd):
            messagebox.showinfo("Success", f"✓ Username changed to: {new_user}")
            show_notification("Username Changed", f"New username: {new_user}")
            self.username = new_user
            self.root.title(f"USB Security Manager Pro - {self.username}")
            for widget in self.sidebar.winfo_children():
                if isinstance(widget, tk.Label) and widget.cget("text") == self.username.upper():
                    widget.config(text=self.username.upper())
            self.new_username.delete(0, tk.END)
            self.user_pwd.delete(0, tk.END)
            self.new_username.focus()
        else:
            messagebox.showerror("Error", "Failed to change username!\nتأكد من صحة كلمة المرور، أو أن الاسم غير مكرر.")

    def change_password_action(self):
        current = self.current_pwd.get()
        new = self.new_pwd.get()
        confirm = self.confirm_pwd.get()
        
        if not current or not new or not confirm:
            messagebox.showwarning("Warning", "Please fill all fields!")
            return
        
        if new != confirm:
            messagebox.showerror("Error", "New passwords do not match!")
            return
            
        from auth_manager import is_strong_password
        if not is_strong_password(new):
            messagebox.showerror("Weak Password", "❌ كلمة المرور ضعيفة جداً!\n\nيجب أن تحتوي على:\n- 8 أحرف على الأقل\n- حرف كبير واحد (A-Z)\n- حرف صغير واحد (a-z)\n- رقم واحد (0-9)")
            return
        
        if current == new:
            messagebox.showerror("Error", "كلمة المرور الجديدة يجب أن تختلف عن الحالية!")
            return
        
        if change_password(self.username, current, new):
            messagebox.showinfo("Success", "✓ Password updated successfully!")
            show_notification("Password Updated", "Your password has been changed successfully")
            self.current_pwd.delete(0, tk.END)
            self.new_pwd.delete(0, tk.END)
            self.confirm_pwd.delete(0, tk.END)
            self.current_pwd.focus()
        else:
            messagebox.showerror("Error", "Failed to change password!\nتأكد من صحة كلمة المرور الحالية.")

    def logout(self):
        if messagebox.askyesno("Confirm Logout", "Are you sure you want to terminate the secure session?"):
            # ── Logout يُعيد البرنامج لحالة "يحتاج Login" ─────────────────
            try:
                from database_manager import set_setting
                set_setting('auto_start_enabled', '0')
            except Exception:
                pass
            disable_startup()
            self.usb_monitor.stop()
            USBBlocker.restore_autoplay()
            if self.tray_icon:
                try: self.tray_icon.stop()
                except Exception: pass
            self.root.destroy()
            # ── فتح شاشة Login من جديد ──────────────────────────────────
            new_root = tk.Tk()
            LoginWindow(new_root)
            new_root.mainloop()

    # ==================== دوال الإجراءات ====================
    def get_connected_usb_devices():
        try:
            import re
            c = wmi.WMI()
            usb_drives = []
            for disk in c.Win32_DiskDrive(InterfaceType="USB"):
                pnp_id = getattr(disk, 'PNPDeviceID', None)
                vid, pid = "N/A", "N/A"
                
                # ✅ استخراج VID/PID من PNPDeviceID
                if pnp_id:
                    vid_m = re.search(r'VID_([0-9A-Fa-f]{4})', pnp_id)
                    pid_m = re.search(r'PID_([0-9A-Fa-f]{4})', pnp_id)
                    vid = vid_m.group(1) if vid_m else "N/A"
                    pid = pid_m.group(1) if pid_m else "N/A"

                device_info = {
                    'model': disk.Model.strip() if disk.Model else "Unknown",
                    'device_id': disk.DeviceID,
                    'pnp_device_id': pnp_id,
                    'serial_number': disk.SerialNumber.strip() if disk.SerialNumber else "N/A",
                    'size_gb': round(int(disk.Size) / (1024 ** 3), 2) if disk.Size else 0,
                    'fingerprint': generate_fingerprint_from_device(disk),
                    'vid': vid,  # ✅ الآن سيتم استخراجها
                    'pid': pid   # ✅ الآن سيتم استخراجها
                }
                usb_drives.append(device_info)
            return usb_drives
        except Exception as e:
            print(f"Error getting USB devices: {e}")
            return []

    def remove_whitelist_device(self):
        """🗑️ إزالة جهاز من القائمة البيضاء -> ينتقل للقائمة السوداء"""
        self._show_devices_for_removal('whitelist', 'blacklist', "Remove from Whitelist")

    def remove_blacklist_device(self):
        """✅ إلغاء حظر جهاز من القائمة السوداء -> ينتقل للقائمة البيضاء"""
        self._show_devices_for_removal('blacklist', 'whitelist', "Unblock from Blacklist")

    def _show_devices_for_removal(self, source_list, target_list, action_name):
        if source_list == 'whitelist':
            devices = get_all_whitelist_devices()
            date_field = 'date_added'
        else:
            devices = get_all_blacklist_devices()
            date_field = 'date_blocked'
        
        if not devices:
            messagebox.showinfo("Info", f"No devices in {source_list}")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title(action_name)
        dialog.configure(bg=Colors.BG_DARK)
        self._center_window(dialog, 750, 550)
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text=f"Select device(s) to {action_name.lower()}:\n(Will move to {'🚫 Blacklist' if target_list=='blacklist' else '✅ Whitelist'})", 
                font=("Segoe UI", 10), fg=Colors.TEXT_MAIN, bg=Colors.BG_DARK, justify=tk.LEFT).pack(pady=15)
        
        frame = tk.Frame(dialog, bg=Colors.BG_CARD)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        headers = ["SELECT", "MODEL", "SERIAL", "SIZE", "ADDED/BLOCKED"]
        for c, h in enumerate(headers):
            tk.Label(frame, text=h, font=("Segoe UI", 9, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_CARD, pady=10).grid(row=0, column=c, sticky='nsew')
        
        selections = {}
        
        for r, dev in enumerate(devices, 1):
            bg_row = Colors.BG_CARD if r % 2 == 0 else Colors.SIDEBAR
            fp = dev.get('fingerprint') or dev.get('id')
            if not fp: continue
            
            var = tk.BooleanVar()
            selections[fp] = var
            
            tk.Checkbutton(frame, variable=var, bg=bg_row, activebackground=bg_row).grid(row=r, column=0, padx=5)
            
            values = [dev.get('model','')[:30], dev.get('serial_number','')[:20], 
                    str(dev.get('size_gb','')) + ' GB' if dev.get('size_gb') else 'N/A',
                    (dev.get(date_field) or '')[:19]]
            for c, val in enumerate(values, start=1):
                tk.Label(frame, text=str(val), fg=Colors.TEXT_MAIN, bg=bg_row).grid(row=r, column=c, sticky='nsew', ipady=5, padx=5)
        
        for i in range(5): frame.grid_columnconfigure(i, weight=1)
        
        def confirm():
            selected = [fp for fp, var in selections.items() if var.get()]
            if not selected:
                messagebox.showwarning("Warning", "⚠️ Please select at least one device")
                return
            if not self.verify_user_password(f"{action_name}\n{len(selected)} device(s) selected"):
                return

            dialog.destroy()
            success_count = 0
            needs_replug  = False
            blocked_now   = False

            for fp in selected:
                result, msg = DeviceListManager.move_device(
                    fp, source_list, target_list,
                    self.username, reason=f"Moved via GUI: {action_name}"
                )
                if result:
                    success_count += 1
                    if msg == "WHITELIST_REPLUG_REQUIRED":
                        needs_replug = True
                    elif msg == "BLACKLIST_BLOCKED":
                        blocked_now = True

            if success_count > 0:
                messagebox.showinfo("Success", f"✓ {success_count}/{len(selected)} device(s) moved successfully!")

                if target_list == 'whitelist' and needs_replug:
                    messagebox.showinfo("🔌 Physical Reconnect Required",
                        "✅ Device moved to Whitelist & Registry cleared.\n\n"
                        "📌 Please UNPLUG the USB drive, wait 3 seconds, then PLUG IT BACK IN.\n"
                        "It will appear in This PC automatically.")
                elif target_list == 'whitelist' and not needs_replug:
                    messagebox.showinfo("✅ Device Enabled",
                        "Device moved to Whitelist and enabled automatically.\n\n"
                        "It should appear in This PC within a few seconds.")
                elif target_list == 'blacklist' and blocked_now:
                    messagebox.showinfo("⛔ Device Blocked",
                        "Device moved to Blacklist and blocked immediately.\n\n"
                        "It has been disabled and will be blocked on every reconnect.")
                elif target_list == 'blacklist' and not blocked_now:
                    messagebox.showinfo("⛔ Added to Blacklist",
                        "Device added to Blacklist.\n\n"
                        "It will be blocked automatically next time it connects.")

                show_notification("Operation Success", f"{success_count} device(s) moved")
                # تحديث كل القوائم المتأثرة
                if target_list == 'whitelist':
                    self.show_whitelist()
                else:
                    self.show_blacklist()
                self.root.after(300, self.show_auto_blocked) if self.current_view == 'auto_blocked' else None
            else:
                messagebox.showerror("Error", "Failed to move devices.")
        
        btn_frame = tk.Frame(dialog, bg=Colors.BG_DARK)
        btn_frame.pack(pady=20)
        
        btn_color = Colors.DANGER if target_list == 'blacklist' else Colors.SUCCESS
        HoverButton(btn_frame, text="✓ Confirm Move", bg=btn_color, fg=Colors.BG_DARK, 
                hover_bg='#cc0044' if target_list == 'blacklist' else '#00cc7a', font=("Segoe UI", 10, "bold"), 
                relief='flat', bd=0, padx=30, command=confirm).pack(side=tk.LEFT, padx=10)
        HoverButton(btn_frame, text="❌ Cancel", bg=Colors.BG_CARD, fg=Colors.TEXT_DIM, 
                hover_bg=Colors.BORDER, font=("Segoe UI", 10), relief='flat', bd=0, padx=30, command=dialog.destroy).pack(side=tk.LEFT, padx=10)
        

    def add_whitelist_device(self):
        """➕ إضافة جهاز للقائمة البيضاء"""
        self._show_source_selector('whitelist')

    def add_blacklist_device(self):
        """⛔ إضافة جهاز للقائمة السوداء"""
        self._show_source_selector('blacklist')

    def _show_source_selector(self, target_list):
        """نافذة اختيار مصدر الجهاز للإضافة"""
        source = tk.Toplevel(self.root)
        source.title("Select Source List")
        source.configure(bg=Colors.BG_DARK)
        self._center_window(source, 420, 320)
        source.transient(self.root)
        source.grab_set()
        
        tk.Label(source, text="📁 Choose source to add from: ", 
                font=("Segoe UI", 11, "bold"), fg=Colors.TEXT_MAIN, bg=Colors.BG_DARK).pack(pady=25)
        
        def on_source_selected(list_name):
            source.destroy()
            self._show_devices_for_adding(list_name, target_list)
        
        sources = [
            ("🚫 From Blacklist", 'blacklist', Colors.DANGER, '#cc0044'),
            ("⏸️ From Auto-Blocked", 'auto_blocked', Colors.WARNING, '#e69500'),
        ]
        
        if target_list == 'blacklist':
            sources.insert(0, ("📋 From Whitelist", 'whitelist', Colors.SUCCESS, '#00cc7a'))
        
        for text, list_name, bg, hover in sources:
            HoverButton(source, text=text, bg=bg, fg=Colors.BG_DARK, hover_bg=hover, 
                         font=("Segoe UI", 10), relief='flat', bd=0, padx=20,
                       command=lambda ln=list_name: on_source_selected(ln)).pack(
                           pady=5, fill=tk.X, padx=30)
        
        HoverButton(source, text="❌ Cancel", bg=Colors.BG_CARD, fg=Colors.TEXT_DIM, 
                   hover_bg=Colors.BORDER, font=("Segoe UI", 10), relief='flat', bd=0, padx=20,
                   command=source.destroy).pack(pady=20, fill=tk.X, padx=30)

    def _show_devices_for_adding(self, source_list, target_list):
        """عرض الأجهزة من قائمة مصدر لإضافتها للقائمة الهدف"""
        
        if source_list == 'blacklist':
            devices = get_all_blacklist_devices()
            date_field = 'date_blocked'
        elif source_list == 'auto_blocked':
            devices = DeviceListManager.get_auto_blocked_devices()
            date_field = 'first_seen'
        elif source_list == 'whitelist':
            devices = get_all_whitelist_devices()
            date_field = 'date_added'
        else:
            messagebox.showerror("Error", f"Unknown source list: {source_list}")
            return
        
        if not devices:
            messagebox.showinfo("Info", f"No devices found in {source_list}")
            return
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Select from {source_list.replace('_', ' ').title()}")
        dialog.configure(bg=Colors.BG_DARK)
        self._center_window(dialog, 750, 550)
        dialog.transient(self.root)
        dialog.grab_set()
        
        target_name = "Whitelist ✅" if target_list == 'whitelist' else "Blacklist 🚫"
        tk.Label(dialog, text=f"Select device(s) to add to {target_name}: ", 
                font=("Segoe UI", 10, "bold"), fg=Colors.TEXT_MAIN, bg=Colors.BG_DARK).pack(pady=15)
        
        frame = tk.Frame(dialog, bg=Colors.BG_CARD)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        headers = ["SELECT", "MODEL", "SERIAL", "SIZE", "DATE"]
        for c, h in enumerate(headers):
            tk.Label(frame, text=h, font=("Segoe UI", 9, "bold"), fg=Colors.PRIMARY, bg=Colors.BG_CARD, pady=10).grid(row=0, column=c, sticky='nsew')
        
        selections = {}
        
        for r, dev in enumerate(devices, 1):
            bg_row = Colors.BG_CARD if r % 2 == 0 else Colors.SIDEBAR
            
            fp = dev.get('fingerprint')
            if not fp or fp == 'None':
                fp = dev.get('id') or dev.get('serial_number')
            
            if not fp:
                continue
            
            var = tk.BooleanVar()
            selections[fp] = var
            
            tk.Checkbutton(frame, variable=var, bg=bg_row, activebackground=bg_row).grid(row=r, column=0, padx=5)
            
            values = [
                dev.get('model','')[:30],
                dev.get('serial_number','')[:20],
                str(dev.get('size_gb','')) + ' GB' if dev.get('size_gb') else 'N/A',
                (dev.get(date_field) or '')[:19]
            ]
            for c, val in enumerate(values, start=1):
                tk.Label(frame, text=str(val), fg=Colors.TEXT_MAIN, bg=bg_row).grid(row=r, column=c, sticky='nsew', ipady=5, padx=5)
        
        for i in range(5): 
            frame.grid_columnconfigure(i, weight=1)
        
        def confirm():
            selected = [fp for fp, var in selections.items() if var.get()]
            
            if not selected:
                messagebox.showwarning("Warning", "⚠️ Please select at least one device")
                return
            
            action_desc = f"Add {len(selected)} device(s) to {target_name}"
            if not self.verify_user_password(action_desc):
                return
            
            dialog.destroy()
            
            success_count = 0
            for fp in selected:
                result, msg = DeviceListManager.move_device(
                    fp, source_list, target_list,
                    self.username,
                    reason=f"Added via GUI from {source_list}"
                )
                if result:
                    success_count += 1
            
            if success_count > 0:
                messagebox.showinfo("Success", f"✓ {success_count}/{len(selected)} device(s) added successfully!")
                show_notification("Device Added", f"{success_count} device(s) added to {target_name}")
                if target_list == 'whitelist':
                    self.show_whitelist()
                else:
                    self.show_blacklist()
            else:
                messagebox.showerror("Error", "Failed to add devices")
        
        btn_frame = tk.Frame(dialog, bg=Colors.BG_DARK)
        btn_frame.pack(pady=20)
        
        btn_color = Colors.SUCCESS if target_list == 'whitelist' else Colors.DANGER
        btn_hover = '#00cc7a' if target_list == 'whitelist' else '#cc0044'
        
        HoverButton(btn_frame, text="✓ Confirm Add", bg=btn_color, fg=Colors.BG_DARK, 
                   hover_bg=btn_hover, font=("Segoe UI", 10, "bold"), 
                   relief='flat', bd=0, padx=30, command=confirm).pack(side=tk.LEFT, padx=10)
        
        HoverButton(btn_frame, text="❌ Cancel", bg=Colors.BG_CARD, fg=Colors.TEXT_DIM, 
                   hover_bg=Colors.BORDER, font=("Segoe UI", 10), 
                   relief='flat', bd=0, padx=30, command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def move_from_auto(self, target_list):
        """نقل جهاز من auto_blocked مع تحديث فوري للواجهة"""
        if not self.verify_user_password(f"Move device to {'Whitelist' if target_list=='whitelist' else 'Blacklist'}"):
            return
            
        selected = [fp for fp, var in self.auto_blocked_selection.items() if var.get()]
        if not selected:
            messagebox.showwarning("Warning", "Please select at least one device")
            return
            
        success_count = 0
        needs_replug  = False
        blocked_now   = False

        for fp in selected:
            result, msg = DeviceListManager.move_device(
                fp, 'auto_blocked', target_list,
                self.username,
                reason="Moved via Auto-Blocked panel"
            )
            if result:
                success_count += 1
                if msg == "WHITELIST_REPLUG_REQUIRED":
                    needs_replug = True
                elif msg == "BLACKLIST_BLOCKED":
                    blocked_now = True

        messagebox.showinfo("Success", f"✓ {success_count}/{len(selected)} device(s) moved successfully!")

        # رسالة مناسبة حسب نتيجة العملية
        if target_list == 'whitelist':
            if needs_replug:
                messagebox.showinfo(
                    "🔌 Reconnect Required",
                    "✅ Device moved to Whitelist.\n\n"
                    "📌 Please UNPLUG and RE-PLUG the USB drive now.\n"
                    "It will be automatically recognized."
                )
            else:
                messagebox.showinfo(
                    "✅ Device Enabled",
                    "Device moved to Whitelist and enabled automatically.\n\n"
                    "It should appear in This PC within a few seconds."
                )
        elif target_list == 'blacklist':
            if blocked_now:
                messagebox.showinfo(
                    "⛔ Device Blocked",
                    "Device moved to Blacklist and blocked immediately.\n\n"
                    "It has been disabled and will be blocked on every reconnect."
                )
            else:
                messagebox.showinfo(
                    "⛔ Added to Blacklist",
                    "Device added to Blacklist.\n\n"
                    "It will be blocked automatically next time it connects."
                )

        show_notification("Move Complete", f"{success_count} device(s) moved from Auto-Blocked")

        # تحديث الواجهة للقائمة المناسبة
        if target_list == 'whitelist':
            self.show_whitelist()
        elif target_list == 'blacklist':
            self.show_blacklist()
        self.root.after(500, self.show_auto_blocked)

# ==================== نقطة الانطلاق ====================

def _run_headless():
    """تشغيل في الخلفية بدون نافذة — يُستدعى من Windows Startup"""
    try:
        from database_manager import get_setting
        username = get_setting('last_username', '').strip() or "admin"
    except Exception:
        username = "admin"
    print(f"🔇 Headless startup as: {username}")
    root = tk.Tk()
    root.withdraw()
    app = MainWindow(root, username, start_hidden=True)
    root.mainloop()

def _run_login():
    """تشغيل نافذة تسجيل الدخول"""
    root = tk.Tk()
    app  = LoginWindow(root)
    root.mainloop()

def _is_auto_start_enabled() -> bool:
    """
    هل البرنامج مُفعَّل للتشغيل التلقائي بدون Login؟
    يُرجع True فقط إذا قام المستخدم بـ Login ناجح مسبقاً
    ولم يقم بـ Logout أو Exit.
    """
    try:
        from database_manager import get_setting
        return get_setting('auto_start_enabled', '0') == '1'
    except Exception:
        return False


def main():
    if not _acquire_mutex():
        try:
            root_tmp = tk.Tk()
            root_tmp.withdraw()
            messagebox.showwarning(
                "USB Shield",
                "⚠️ USB Shield is already running.\n\n"
                "Check the system tray icon near the clock."
            )
            root_tmp.destroy()
        except Exception:
            pass
        sys.exit(0)

    if "--startup" in sys.argv and _is_auto_start_enabled():
        # ── تشغيل تلقائي من Windows Startup بعد Login ناجح سابق ─────────
        # بدون نافذة، بدون CMD، مباشرة في الخلفية
        _run_headless()
    else:
        # ── تشغيل عادي أو بعد Logout/Exit → يطلب Login ─────────────────
        _run_login()

if __name__ == "__main__":
    main()